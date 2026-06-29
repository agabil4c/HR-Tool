import calendar as month_calendar
import csv
import json
import logging
import os
import re
import smtplib
import ssl
from email.message import EmailMessage
from io import StringIO
from io import BytesIO
from datetime import date, datetime, timedelta
from typing import Any

from sqlalchemy import func, text
from sqlalchemy.orm import Session

from . import models


STAFF_ROLE_ALIASES = {"staff", "employee"}
FALLBACK_PRIORITY = ["UG", "RW", "BI"]
DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS = 21
DEFAULT_ALLOWED_SECTIONS_BY_ROLE = {
    "super-admin": ["*"],
    "hr-manager": ["dashboard", "profile", "staff", "departments", "leave", "attendance", "reports", "calendar", "user-admin"],
    "hr-officer": ["dashboard", "profile", "staff", "departments", "leave", "attendance", "reports", "calendar"],
    "department-manager": ["dashboard", "profile", "staff", "departments", "leave", "attendance", "reports", "calendar"],
    "line-manager": ["dashboard", "profile", "staff", "departments", "leave", "attendance", "reports", "calendar"],
    "dept-head": ["dashboard", "profile", "staff", "departments", "leave", "attendance", "reports", "calendar"],
    "department-head": ["dashboard", "profile", "staff", "departments", "leave", "attendance", "reports", "calendar"],
    "general-manager": ["dashboard", "profile", "staff", "departments", "leave", "attendance", "reports", "calendar"],
    "staff": ["dashboard", "profile", "leave", "attendance", "reports", "calendar"],
    "employee": ["dashboard", "profile", "leave", "attendance", "reports", "calendar"],
}

logger = logging.getLogger("uvicorn.error")

MANAGER_DASHBOARD_ROUTE = "/dashboard/managers/index"


def ensure_employee_extended_fields(db: Session) -> None:
    existing_columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(employees)")).fetchall()
    }

    alter_statements = {
        "salary_benefits": "ALTER TABLE employees ADD COLUMN salary_benefits TEXT",
        "bank_account": "ALTER TABLE employees ADD COLUMN bank_account TEXT",
        "account_names": "ALTER TABLE employees ADD COLUMN account_names TEXT",
        "bank_name": "ALTER TABLE employees ADD COLUMN bank_name TEXT",
    }

    did_change = False
    for column_name, statement in alter_statements.items():
        if column_name not in existing_columns:
            db.execute(text(statement))
            did_change = True

    if did_change:
        db.commit()


def ensure_attendance_extended_fields(db: Session) -> None:
    existing_columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(attendance_records)")).fetchall()
    }

    alter_statements = {
        "clock_in": "ALTER TABLE attendance_records ADD COLUMN clock_in TEXT",
        "clock_out": "ALTER TABLE attendance_records ADD COLUMN clock_out TEXT",
        "total_hours": "ALTER TABLE attendance_records ADD COLUMN total_hours TEXT",
    }

    did_change = False
    for column_name, statement in alter_statements.items():
        if column_name not in existing_columns:
            db.execute(text(statement))
            did_change = True

    if did_change:
        db.commit()


def ensure_leave_handover_fields(db: Session) -> None:
    existing_columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(leave_applications)")).fetchall()
    }

    alter_statements = {
        "handover_report": "ALTER TABLE leave_applications ADD COLUMN handover_report TEXT",
        "handover_submitted_on": "ALTER TABLE leave_applications ADD COLUMN handover_submitted_on TEXT",
    }

    did_change = False
    for column_name, statement in alter_statements.items():
        if column_name not in existing_columns:
            db.execute(text(statement))
            did_change = True

    if did_change:
        db.commit()


def _normalized_role(role: str | None) -> str:
    import re
    return re.sub(r'[\s_]+', '-', (role or "").strip().lower())


def _is_staff_role(role: str | None) -> bool:
    return _normalized_role(role) in STAFF_ROLE_ALIASES


def _extract_section_keys_from_permissions(permissions: list[str]) -> list[str]:
    """Extract section keys from granular permissions.
    
    Maps permissions like ['staff.view', 'staff.create', 'dashboard.view', 'roles.manage']
    to section keys like ['staff', 'dashboard', 'role-access'] for authorization checks.
    """
    section_map = {
        'dashboard': 'dashboard',
        'profile': 'profile',
        'staff': 'staff',
        'departments': 'departments',
        'leave': 'leave',
        'attendance': 'attendance',
        'reports': 'reports',
        'calendar': 'calendar',
        'users': 'user-admin',
        'roles': 'role-access',
    }
    
    sections = set()
    for permission in (permissions or []):
        # Extract the prefix before the dot, e.g., 'staff.view' -> 'staff'
        section = permission.split('.')[0] if '.' in permission else permission
        mapped_section = section_map.get(section, section)
        if mapped_section:
            sections.add(mapped_section)
    
    return sorted(list(sections))


def _can_view_cross_department(user: models.User) -> bool:
    return _normalized_role(user.role) in {"super-admin", "hr-manager"}


def _get_department_for_user(db: Session, user: models.User) -> models.Department | None:
    if not user.department:
        return None
    return (
        db.query(models.Department)
        .filter(func.lower(models.Department.name) == user.department.strip().lower())
        .first()
    )


def _get_scoped_employee_query(db: Session, user: models.User):
    query = db.query(models.Employee)
    if _can_view_cross_department(user):
        return query

    department = _get_department_for_user(db, user)
    if not department:
        return query.filter(models.Employee.id == -1)

    return query.filter(models.Employee.department_id == department.id)


def _get_employee_for_user(db: Session, user: models.User) -> models.Employee | None:
    normalized_name = (user.full_name or "").strip().lower()
    scoped_query = _get_scoped_employee_query(db, user)

    if normalized_name:
        direct_match = scoped_query.filter(func.lower(models.Employee.name) == normalized_name).first()
        if direct_match:
            return direct_match

    if _is_staff_role(user.role):
        return None

    return scoped_query.first()


def resolve_dashboard_route(
    role: str,
    department: str | None = None,
    access_level: str | None = None,
    configured_route: str | None = None,
) -> str:
    # Always calculate based on role to ensure consistency
    # Ignore potentially stale configured_route values
    normalized_role = _normalized_role(role)
    if normalized_role == "super-admin":
        return "/index"
    if normalized_role == "hr-manager":
        return "/hr"
    if normalized_role in {"department-manager", "line-manager", "dept-head", "department-head", "general-manager"}:
        return MANAGER_DASHBOARD_ROUTE
    if normalized_role in STAFF_ROLE_ALIASES:
        return "/dashboard"
    return "/dashboard"


def build_user_session_payload(user: models.User) -> dict[str, Any]:
    normalized_role = _normalized_role(user.role)
    allowed_sections = ["*"] if normalized_role == "super-admin" else (user.allowed_sections or [])
    baseline_sections = DEFAULT_ALLOWED_SECTIONS_BY_ROLE.get(normalized_role, DEFAULT_ALLOWED_SECTIONS_BY_ROLE.get("employee", []))
    if not allowed_sections:
        allowed_sections = baseline_sections
    else:
        allowed_sections = sorted(set([*allowed_sections, *baseline_sections]))

    normalized_access_level = _normalized_role(user.access_level)
    if normalized_access_level in {"", "self-service"} and normalized_role in {"department-manager", "line-manager", "dept-head", "department-head", "general-manager"}:
        resolved_access_level = "department-manager"
    elif normalized_access_level in {"", "self-service"} and normalized_role in {"staff", "employee"}:
        resolved_access_level = "employee"
    else:
        resolved_access_level = normalized_access_level or normalized_role or "employee"

    return {
        "role": user.role,
        "fullName": user.full_name,
        "department": user.department or "General",
        "accessLevel": resolved_access_level,
        "dashboardRoute": resolve_dashboard_route(
            user.role,
            user.department,
            user.access_level,
            user.dashboard_route,
        ),
        "allowedSections": allowed_sections,
    }


def _department_map(db: Session) -> dict[int, str]:
    return {d.id: d.name for d in db.query(models.Department).all()}


def get_module_data(db: Session, module_key: str, user: models.User, year: int | None = None, month: int | None = None) -> dict[str, Any] | None:
    ensure_employee_extended_fields(db)
    ensure_attendance_extended_fields(db)
    if module_key == "hr-dashboard":
        return _get_hr_dashboard(db, user)
    if module_key == "staff-biodata":
        return _get_staff_biodata(db, user)
    if module_key == "department":
        return _get_department_module(db, user)
    if module_key == "holidays":
        return _get_holidays_module(db)
    if module_key == "attendance":
        return _get_attendance_module(db, user)
    if module_key == "attendance-main":
        return _get_attendance_main_module(db, user)
    if module_key == "performance-reviews":
        return _get_performance_reviews_module(db, user)
    if module_key == "dashboard":
        return _get_my_profile_module(db, user, year=year, month=month)
    if module_key == "my-bio":
        return _get_my_bio_module(db, user)
    if module_key == "leave-requests":
        return _get_leave_requests_module(db, user)
    if module_key == "calendar":
        return _get_calendar_module(db)
    return None


def _get_hr_dashboard(db: Session, user: models.User) -> dict[str, Any]:
    if _can_view_cross_department(user):
        departments = db.query(models.Department).all()
    else:
        department = _get_department_for_user(db, user)
        departments = [department] if department else []

    scoped_employees = _get_scoped_employee_query(db, user)
    employee_count = scoped_employees.count()
    review_count = db.query(models.PerformanceReview).count()
    template_count = db.query(models.ReviewTemplate).count()
    approval_count = db.query(models.ApprovalRequest).filter(models.ApprovalRequest.status == "Pending").count()

    stat_cards = [
        {
            "title": "Active Staff",
            "value": f"{employee_count:,}",
            "icon": "solar:users-group-rounded-bold-duotone",
            "tone": "text-primary bg-primary/10",
        },
        {
            "title": "Departments",
            "value": str(len(departments)),
            "icon": "solar:buildings-3-bold-duotone",
            "tone": "text-warning bg-warning/10",
        },
        {
            "title": "Active Form Templates",
            "value": str(template_count),
            "icon": "solar:clipboard-list-bold-duotone",
            "tone": "text-success bg-success/10",
        },
        {
            "title": "Reviews In Progress",
            "value": str(review_count),
            "icon": "solar:document-text-bold-duotone",
            "tone": "text-violet-500 bg-violet-500/10",
        },
    ]

    departments_payload = [
        {
            "name": d.name,
            "head": d.head or "N/A",
            "members": f"{d.staff_count} Staff Members",
            "icon": d.icon or "solar:buildings-3-bold-duotone",
            "tone": d.icon_bg or "text-primary bg-primary/10",
        }
        for d in departments
    ]

    templates = [
        {
            "name": t.name,
            "audience": t.audience,
            "age": t.age,
            "tag": t.tag,
            "icon": t.icon,
            "tone": t.tone,
        }
        for t in db.query(models.ReviewTemplate).all()
    ]

    pending_bio = [
        {
            "name": e.name,
            "status": "READY FOR REVIEW" if e.status == "Active" else "MISSING ID VERIFICATION",
            "avatar": e.avatar or "avatar2",
            "cta": "Verify",
        }
        for e in scoped_employees.limit(5).all()
    ]

    approval_queue = [
        {
            "id": a.id,
            "employee": a.employee,
            "period": a.period,
            "approver": a.approver,
            "status": a.status,
        }
        for a in db.query(models.ApprovalRequest).all()
    ]

    line_manager_stats = [
        {
            "title": "Pending Approvals",
            "value": str(approval_count),
            "subtitle": f"{approval_count} awaiting action",
            "icon": "solar:clipboard-check-outline",
            "accent": "text-primary",
            "subtitleTone": "text-success",
        },
        {
            "title": "Team Attendance",
            "value": "94%",
            "subtitle": "Live attendance snapshot",
            "icon": "solar:users-group-rounded-outline",
            "accent": "text-orange-500",
            "subtitleTone": "text-default-500",
        },
        {
            "title": "Performance Reviews",
            "value": f"{review_count} Due",
            "subtitle": "Current cycle",
            "icon": "solar:star-outline",
            "accent": "text-violet-500",
            "subtitleTone": "text-default-500",
        },
    ]

    today = date.today()
    current_week_start = today - timedelta(days=today.weekday())
    current_week_label = current_week_start.isoformat()
    current_week_columns = [
        (current_week_start + timedelta(days=offset)).strftime("%a %d")
        for offset in range(5)
    ]

    roster_rows = [
        {
            "initials": r.initials or "--",
            "name": r.employee_name,
            "statuses": r.statuses or [],
        }
        for r in (
            db.query(models.RosterEntry)
            .filter(models.RosterEntry.week_label == current_week_label)
            .all()
        )
    ]

    if not roster_rows:
        current_week_columns = []

    return {
        "statCards": stat_cards,
        "departments": departments_payload,
        "quickActions": [
            {"title": "New Employee User", "subtitle": "Assign roles and permissions", "icon": "solar:user-plus-bold-duotone"},
            {"title": "Bio Data Entry", "subtitle": "Collect personal documentation", "icon": "solar:document-add-bold-duotone"},
            {"title": "Bulk User Upload", "subtitle": "Import CSV employee list", "icon": "solar:upload-bold-duotone"},
        ],
        "reviewTemplates": templates,
        "pendingBio": pending_bio,
        "lineManagerStats": line_manager_stats,
        "approvalQueue": approval_queue,
        "rosterColumns": current_week_columns,
        "rosterRows": roster_rows,
    }


def _get_staff_biodata(db: Session, user: models.User) -> dict[str, Any]:
    dept_map = _department_map(db)
    scoped_employees = _get_scoped_employee_query(db, user)

    if _can_view_cross_department(user):
        departments = db.query(models.Department).all()
    else:
        department = _get_department_for_user(db, user)
        departments = [department] if department else []

    staff_profiles: list[dict[str, Any]] = []
    for e in scoped_employees.all():
        linked_user = None
        if e.user_id:
            linked_user = db.query(models.User).filter(models.User.id == e.user_id).first()

        if linked_user:
            leave_balance = get_user_leave_balance_summary(db, linked_user)
        else:
            default_initial = _get_department_initial_leave_days_by_name(db, dept_map.get(e.department_id, ""))
            leave_balance = {
                "initialDays": default_initial,
                "usedDays": 0,
                "remainingDays": default_initial,
            }

        staff_profiles.append(
            {
                "initials": e.initials,
                "name": e.name,
                "dbId": e.id,
                "firstName": e.first_name,
                "lastName": e.last_name,
                "gender": e.gender,
                "dateOfBirth": e.date_of_birth,
                "nationality": e.nationality,
                "maritalStatus": e.marital_status,
                "profilePhoto": e.profile_photo,
                "nationalId": e.national_id,
                "personalEmail": e.personal_email,
                "workEmail": e.work_email,
                "phone": e.contact,
                "emergencyContact": {
                    "name": e.emergency_contact_name,
                    "phone": e.emergency_contact_phone,
                    "relationship": e.emergency_contact_relationship,
                },
                "address": {
                    "city": e.address_city,
                    "district": e.address_district,
                    "country": e.address_country,
                    "line1": e.address_line1,
                },
                "employeeId": e.employee_code,
                "startedAt": e.started_at,
                "role": e.role,
                "dept": dept_map.get(e.department_id, "UNASSIGNED").upper(),
                "reporting": e.reporting or [],
                "status": e.status,
                "employmentType": e.employment_type,
                "dateOfJoining": e.date_of_joining,
                "workLocation": e.work_location,
                "salary": e.salary,
                "payGrade": e.pay_grade,
                "salaryBenefits": e.salary_benefits,
                "bankAccount": e.bank_account,
                "accountNames": e.account_names,
                "bankName": e.bank_name,
                "bankDetails": e.bank_details,
                "taxId": e.tax_id,
                "nssfNumber": e.nssf_number,
                "leaveBalance": leave_balance,
                "documents": {
                    "cv": e.cv_document,
                    "contract": e.contract_document,
                    "idCopy": e.id_copy_document,
                    "certificates": e.certificates_document,
                    "other": e.other_documents,
                },
            }
        )

    approval_cards = [
        {
            "dept": f"{d.name} Dept",
            "levels": d.approval_level.replace("-Step ", " LEVELS") if "Step" in d.approval_level else d.approval_level,
            "head": f"Dept Head: {d.head or 'N/A'}",
            "count": f"+{d.staff_count}",
        }
        for d in departments
    ]

    return {
        "sidebarItems": [
            {"label": "Directory", "icon": "solar:users-group-rounded-outline", "active": True},
            {"label": "Org Chart", "icon": "solar:sitemap-outline", "active": False},
            {"label": "Onboarding", "icon": "solar:file-text-outline", "active": False},
            {"label": "Performance", "icon": "solar:graph-up-outline", "active": False},
        ],
        "hierarchyItems": [
            {"label": "Approval Flows", "icon": "solar:shield-check-outline"},
            {"label": "Assign Managers", "icon": "solar:user-check-outline"},
        ],
        "staffProfiles": staff_profiles,
        "approvalCards": approval_cards,
    }


def _get_department_module(db: Session, user: models.User) -> dict[str, Any]:
    if _can_view_cross_department(user):
        departments = db.query(models.Department).all()
    else:
        department = _get_department_for_user(db, user)
        departments = [department] if department else []

    leave_policy_map = _get_department_leave_policy_map(db)

    rows = [
        {
            "iconBg": d.icon_bg or "bg-primary/10 text-primary",
            "icon": d.icon or "solar:buildings-3-bold-duotone",
            "name": d.name,
            "head": d.head or "N/A",
            "avatarColor": d.avatar_color or "bg-slate-200 dark:bg-slate-700",
            "staff": f"{d.staff_count} Employees",
            "approvalLevel": d.approval_level,
            "status": d.status,
            "initialLeaveDays": int(leave_policy_map.get(d.id, DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS)),
        }
        for d in departments
    ]
    return {"departmentRows": rows}


def _get_holidays_module(db: Session) -> dict[str, Any]:
    records: list[dict[str, Any]] = []

    for holiday in db.query(models.Holiday).all():
        holiday_date = _parse_holiday_date(holiday.date)
        records.append(
            {
                "recordKey": f"holiday-{holiday.id}",
                "recordId": holiday.id,
                "sourceType": "holiday",
                "day": holiday.day,
                "date": holiday.date,
                "rawDate": holiday_date.isoformat() if holiday_date else "",
                "name": holiday.name,
                "type": holiday.type,
                "kind": "Holiday",
            }
        )

    for event in db.query(models.CalendarEvent).all():
        if event.title and event.title.strip().endswith(" Calendar"):
            parts = event.title.strip().split(" ")
            if len(parts) == 3 and parts[0].isdigit() and len(parts[0]) == 4:
                continue

        event_date = _parse_calendar_event_date(event.start_date)
        if not event_date:
            continue

        records.append(
            {
                "recordKey": f"event-{event.id}",
                "recordId": event.id,
                "sourceType": "event",
                "day": event_date.strftime("%A"),
                "date": _format_holiday_display_date(event_date),
                "rawDate": event_date.isoformat(),
                "name": event.title,
                "type": _event_category_from_class_names(event.class_names or []),
                "kind": "Event",
            }
        )

    records.sort(key=lambda item: (item.get("rawDate") or "", item.get("name") or ""))

    return {
        "holidays": records
    }


def _parse_holiday_date(value: str | None) -> date | None:
    if not value:
        return None

    raw = value.strip()
    if not raw:
        return None

    for fmt in ("%Y-%m-%d", "%d %b", "%d %B", "%d %b %Y", "%d %B %Y"):
        try:
            parsed = datetime.strptime(raw, fmt)
            if fmt in {"%d %b", "%d %B"}:
                return date(date.today().year, parsed.month, parsed.day)
            return parsed.date()
        except ValueError:
            continue

    return None


def _format_holiday_display_date(holiday_date: date) -> str:
    return holiday_date.strftime("%d %b %Y")


def _parse_calendar_event_date(value: str | None) -> date | None:
    if not value:
        return None

    raw = value.strip()
    if not raw:
        return None

    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(raw.replace("Z", "")).date()
    except ValueError:
        return None


def _serialize_calendar_event(row: models.CalendarEvent) -> dict[str, Any]:
    return {
        "id": f"event-{row.id}",
        "eventId": row.id,
        "title": row.title,
        "start": row.start_date,
        "classNames": row.class_names or [],
        "extendedProps": {
            "sourceType": "event",
            "readOnly": False,
        },
    }


def _event_category_from_class_names(class_names: list[str]) -> str:
    class_name = (class_names or ["!text-primary"])[0]
    mapping = {
        "!text-primary": "Meeting",
        "!text-success": "Team Building",
        "!text-info": "Workshop",
        "!text-warning": "Town Hall",
        "!text-danger": "Urgent Event",
    }
    return mapping.get(class_name, "Calendar Event")


def _event_class_from_category(category: str | None) -> str:
    mapping = {
        "Meeting": "!text-primary",
        "Team Building": "!text-success",
        "Workshop": "!text-info",
        "Town Hall": "!text-warning",
        "Urgent Event": "!text-danger",
    }
    return mapping.get((category or "").strip(), "!text-primary")


def _serialize_holiday_event(row: models.Holiday) -> dict[str, Any] | None:
    holiday_date = _parse_holiday_date(row.date)
    if not holiday_date:
        return None

    return {
        "id": f"holiday-{row.id}",
        "holidayId": row.id,
        "title": row.name,
        "start": holiday_date.isoformat(),
        "classNames": ["!text-danger"],
        "extendedProps": {
            "sourceType": "holiday",
            "holidayType": row.type,
            "readOnly": True,
        },
    }


def create_public_holiday(db: Session, payload: dict[str, Any]) -> dict[str, Any]:
    holiday_date = payload["date"]
    row = models.Holiday(
        day=holiday_date.strftime("%A"),
        date=_format_holiday_display_date(holiday_date),
        name=payload["name"].strip(),
        type=payload.get("type", "Gazetted Holiday").strip() or "Gazetted Holiday",
    )

    db.add(row)
    db.commit()
    db.refresh(row)

    return {
        "id": row.id,
        "day": row.day,
        "date": row.date,
        "name": row.name,
        "type": row.type,
    }


def update_public_holiday(db: Session, holiday_id: int, payload: dict[str, Any]) -> dict[str, Any] | None:
    row = db.query(models.Holiday).filter(models.Holiday.id == holiday_id).first()
    if not row:
        return None

    holiday_date = payload["date"]
    row.day = holiday_date.strftime("%A")
    row.date = _format_holiday_display_date(holiday_date)
    row.name = payload["name"].strip()
    row.type = payload.get("type", row.type).strip() or row.type
    db.commit()
    db.refresh(row)

    return {
        "id": row.id,
        "day": row.day,
        "date": row.date,
        "name": row.name,
        "type": row.type,
    }


def delete_public_holiday(db: Session, holiday_id: int) -> bool:
    row = db.query(models.Holiday).filter(models.Holiday.id == holiday_id).first()
    if not row:
        return False

    db.delete(row)
    db.commit()
    return True


def _global_annual_observances(year: int) -> list[tuple[date, str]]:
    return [
        (date(year, 1, 1), "New Year's Day"),
        (date(year, 5, 1), "International Workers' Day"),
        (date(year, 12, 25), "Christmas Day"),
    ]


def sync_annual_public_holidays(
    db: Session,
    year: int,
    country_code: str,
    region_code: str | None = None,
    include_global: bool = True,
) -> dict[str, Any]:
    try:
        import holidays as holidays_lib
    except ImportError as exc:
        raise RuntimeError("python-holidays package is not installed") from exc

    normalized_country = (country_code or "").strip().upper()
    if not normalized_country:
        raise ValueError("countryCode is required")

    normalized_region = (region_code or "").strip().upper() or None
    supported_countries = {str(code).upper() for code in holidays_lib.list_supported_countries().keys()}
    logger.warning(
        "python-holidays supported country codes count=%s countries=%s",
        len(supported_countries),
        ",".join(sorted(supported_countries)),
    )

    fallback_candidates = [code for code in FALLBACK_PRIORITY if code in supported_countries]
    if not fallback_candidates:
        raise RuntimeError(
            f"None of fallback countries are supported by python-holidays: {', '.join(FALLBACK_PRIORITY)}"
        )

    country_aliases = {
        "UK": "GB",
    }
    requested_country = normalized_country
    normalized_country = country_aliases.get(normalized_country, normalized_country)
    if normalized_country not in supported_countries:
        normalized_country = fallback_candidates[0]
        normalized_region = None

    existing_pairs = {
        ((_parse_holiday_date(row.date) or date.min).isoformat(), (row.name or "").strip().lower())
        for row in db.query(models.Holiday).all()
    }

    created = 0

    if include_global:
        for holiday_date, holiday_name in _global_annual_observances(year):
            key = (holiday_date.isoformat(), holiday_name.strip().lower())
            if key in existing_pairs:
                continue

            db.add(
                models.Holiday(
                    day=holiday_date.strftime("%A"),
                    date=_format_holiday_display_date(holiday_date),
                    name=holiday_name,
                    type="Global Annual Holiday",
                )
            )
            existing_pairs.add(key)
            created += 1

    annual_holidays = None
    lookup_errors: list[str] = []
    countries_to_try = [normalized_country] + [
        code for code in fallback_candidates if code != normalized_country
    ]

    for candidate_country in countries_to_try:
        candidate_region = normalized_region if candidate_country == normalized_country else None
        try:
            annual_holidays = holidays_lib.country_holidays(
                candidate_country,
                years=[year],
                subdiv=candidate_region,
            )
            normalized_country = candidate_country
            normalized_region = candidate_region
            break
        except Exception as exc:
            # If subdivision is invalid for this country, retry with country-wide holidays.
            if candidate_region:
                try:
                    annual_holidays = holidays_lib.country_holidays(
                        candidate_country,
                        years=[year],
                        subdiv=None,
                    )
                    normalized_country = candidate_country
                    normalized_region = None
                    break
                except Exception as regionless_exc:
                    lookup_errors.append(f"{candidate_country}(regionless): {regionless_exc}")
            lookup_errors.append(f"{candidate_country}: {exc}")

    if annual_holidays is None:
        raise ValueError(
            "Invalid countryCode or regionCode for annual holiday lookup. "
            + "Fallback attempts: "
            + " | ".join(lookup_errors)
        )

    region_suffix = f" ({normalized_region})" if normalized_region else ""
    holiday_type = f"{normalized_country}{region_suffix} Annual Holiday"

    for holiday_date, holiday_name in annual_holidays.items():
        key = (holiday_date.isoformat(), str(holiday_name).strip().lower())
        if key in existing_pairs:
            continue

        db.add(
            models.Holiday(
                day=holiday_date.strftime("%A"),
                date=_format_holiday_display_date(holiday_date),
                name=str(holiday_name),
                type=holiday_type,
            )
        )
        existing_pairs.add(key)
        created += 1

    db.commit()

    return {
        "year": year,
        "requestedCountryCode": requested_country,
        "countryCode": normalized_country,
        "regionCode": normalized_region,
        "includeGlobal": include_global,
        "created": created,
    }


def create_calendar_event(db: Session, payload: dict[str, Any]) -> dict[str, Any]:
    row = models.CalendarEvent(
        title=payload["title"].strip(),
        start_date=payload["startDate"],
        class_names=payload.get("classNames", []),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _serialize_calendar_event(row)


def update_calendar_event(db: Session, event_id: int, payload: dict[str, Any]) -> dict[str, Any] | None:
    row = db.query(models.CalendarEvent).filter(models.CalendarEvent.id == event_id).first()
    if not row:
        return None

    row.title = payload["title"].strip()
    if payload.get("startDate"):
        row.start_date = payload["startDate"]
    row.class_names = payload.get("classNames", [])
    db.commit()
    db.refresh(row)
    return _serialize_calendar_event(row)


def delete_calendar_event(db: Session, event_id: int) -> bool:
    row = db.query(models.CalendarEvent).filter(models.CalendarEvent.id == event_id).first()
    if not row:
        return False

    db.delete(row)
    db.commit()
    return True


def get_upcoming_events_summary(db: Session, limit: int = 8) -> list[dict[str, Any]]:
    today = date.today()
    entries: list[dict[str, Any]] = []

    for holiday in db.query(models.Holiday).all():
        holiday_date = _parse_holiday_date(holiday.date)
        if not holiday_date or holiday_date < today:
            continue

        entries.append(
            {
                "id": f"holiday-{holiday.id}",
                "title": holiday.name,
                "date": holiday_date.isoformat(),
                "kind": "Holiday",
                "type": holiday.type,
            }
        )

    for event in db.query(models.CalendarEvent).all():
        # Skip system-seeded month placeholders (e.g. "2026 January Calendar").
        if event.title and event.title.strip().endswith(" Calendar"):
            parts = event.title.strip().split(" ")
            if len(parts) == 3 and parts[0].isdigit() and len(parts[0]) == 4:
                continue

        event_date = _parse_calendar_event_date(event.start_date)
        if not event_date or event_date < today:
            continue

        entries.append(
            {
                "id": f"event-{event.id}",
                "title": event.title,
                "date": event_date.isoformat(),
                "kind": "Event",
                "type": "Calendar Event",
            }
        )

    entries.sort(key=lambda item: (item["date"], item["title"].lower()))
    return entries[:limit]


def _get_attendance_module(db: Session, user: models.User) -> dict[str, Any]:
    current_employee = _get_employee_for_user(db, user)
    if _is_staff_role(user.role):
        employees = [current_employee] if current_employee else []
    else:
        employees = _get_scoped_employee_query(db, user).all()

    employee_ids = [employee.id for employee in employees]
    if not employee_ids:
        records = []
    else:
        records = (
            db.query(models.AttendanceRecord)
            .filter(models.AttendanceRecord.employee_id.in_(employee_ids))
            .all()
        )

    employee_by_id = {employee.id: employee for employee in employees}

    approved_hours = sum(r.work_hours_val for r in records if r.approval_status == "Approved")
    rejected_hours = sum(r.work_hours_val for r in records if r.approval_status == "Rejected")
    pending_hours = sum(r.work_hours_val for r in records if r.approval_status == "Pending")

    work_details = [
        {
            "id": 1,
            "value": int(approved_hours),
            "label": "Approved Hours",
            "icon": "clock",
            "textColor": "text-info",
            "bgColor": "bg-info/10",
        },
        {
            "id": 2,
            "value": int(rejected_hours),
            "label": "Rejected Hours",
            "icon": "rejected",
            "textColor": "text-danger",
            "bgColor": "bg-danger/10",
        },
        {
            "id": 3,
            "value": int(pending_hours),
            "label": "Pending Hours",
            "icon": "pending",
            "textColor": "text-warning",
            "bgColor": "bg-warning/10",
        },
    ]

    rows = [
        {
            "employeeId": r.employee_id,
            "employeeName": (employee_by_id.get(r.employee_id).name if employee_by_id.get(r.employee_id) else "Unknown Employee"),
            "date": r.date,
            "day": r.day,
            "checkIn": r.check_in,
            "checkOut": r.check_out,
            "clockIn": r.clock_in,
            "clockOut": r.clock_out,
            "mealBreak": r.meal_break,
            "totalHours": r.total_hours,
            "workHours": r.work_hours,
            "overtime": r.overtime,
            "approvalStatus": r.approval_status,
            "isCurrentUser": bool(current_employee and r.employee_id == current_employee.id),
        }
        for r in records
    ]

    employees_payload = [
        {
            "id": employee.id,
            "name": employee.name,
            "role": employee.role or "",
            "status": employee.status or "Active",
            "employeeCode": employee.employee_code or "",
            "dateOfJoining": employee.date_of_joining or "",
            "department": user.department or "",
            "isCurrentUser": bool(current_employee and employee.id == current_employee.id),
        }
        for employee in employees
    ]

    return {
        "workDetails": work_details,
        "records": rows,
        "employees": employees_payload,
        "currentEmployeeId": current_employee.id if current_employee else None,
    }


def _normalize_attendance_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%I:%M %p")
    if isinstance(value, date):
        return value.strftime("%d %b %Y")
    return str(value).strip()


def _parse_excel_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = str(value).strip()
    if not raw:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%d %b %Y",
        "%d %B %Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue

    return None


def _to_hours_value(value: Any) -> float:
    cleaned = re.sub(r"[^0-9.]", "", str(value or ""))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def import_attendance_from_excel(
    db: Session,
    *,
    employee: models.Employee,
    from_date: date,
    to_date: date,
    file_bytes: bytes,
    file_name: str | None = None,
) -> dict[str, Any]:
    ensure_attendance_extended_fields(db)
    lowered_name = (file_name or "").strip().lower()
    if lowered_name.endswith(".csv"):
        try:
            decoded = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            decoded = file_bytes.decode("latin-1")
        rows = [tuple(row) for row in csv.reader(StringIO(decoded))]
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("openpyxl is required to import Excel files") from exc

        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise ValueError("The uploaded Excel file is empty")

    header_row = rows[0]
    header_map = {
        _normalize_attendance_header(cell): idx
        for idx, cell in enumerate(header_row)
        if _normalize_attendance_header(cell)
    }

    def _find_col(*keys: str) -> int | None:
        for key in keys:
            idx = header_map.get(key)
            if idx is not None:
                return idx
        return None

    date_idx = _find_col("date", "attendancedate", "workdate")
    if date_idx is None:
        raise ValueError("Excel must include a 'Date' column")

    day_idx = _find_col("day", "weekday")
    check_in_idx = _find_col("checkin")
    check_out_idx = _find_col("checkout")
    clock_in_idx = _find_col("clockin", "timein")
    clock_out_idx = _find_col("clockout", "timeout")
    meal_break_idx = _find_col("mealbreak", "break", "lunchbreak")
    total_hours_idx = _find_col("totalhours")
    work_hours_idx = _find_col("workedhours", "workhours", "hoursworked")
    overtime_idx = _find_col("overtime", "ot")
    status_idx = _find_col("status", "approvalstatus", "approval")

    existing_records = (
        db.query(models.AttendanceRecord)
        .filter(models.AttendanceRecord.employee_id == employee.id)
        .all()
    )
    replaced_count = 0
    for record in existing_records:
        parsed = _parse_holiday_date(record.date)
        if parsed and from_date <= parsed <= to_date:
            db.delete(record)
            replaced_count += 1

    inserted_count = 0
    skipped_count = 0

    for row in rows[1:]:
        raw_date = row[date_idx] if date_idx < len(row) else None
        parsed_date = _parse_excel_date(raw_date)
        if not parsed_date:
            skipped_count += 1
            continue
        if parsed_date < from_date or parsed_date > to_date:
            continue

        scheduled_check_in = _to_string(row[check_in_idx]) if check_in_idx is not None and check_in_idx < len(row) else ""
        scheduled_check_out = _to_string(row[check_out_idx]) if check_out_idx is not None and check_out_idx < len(row) else ""
        actual_clock_in = _to_string(row[clock_in_idx]) if clock_in_idx is not None and clock_in_idx < len(row) else ""
        actual_clock_out = _to_string(row[clock_out_idx]) if clock_out_idx is not None and clock_out_idx < len(row) else ""

        total_hours_value = _to_string(row[total_hours_idx]) if total_hours_idx is not None and total_hours_idx < len(row) else ""
        day_value = _to_string(row[day_idx]) if day_idx is not None and day_idx < len(row) else parsed_date.strftime("%a")
        work_hours_value = _to_string(row[work_hours_idx]) if work_hours_idx is not None and work_hours_idx < len(row) else ""
        status_value = _to_string(row[status_idx]) if status_idx is not None and status_idx < len(row) else "Approved"

        if not scheduled_check_in:
            scheduled_check_in = actual_clock_in
        if not scheduled_check_out:
            scheduled_check_out = actual_clock_out
        if not actual_clock_in:
            actual_clock_in = scheduled_check_in
        if not actual_clock_out:
            actual_clock_out = scheduled_check_out
        if not total_hours_value:
            total_hours_value = work_hours_value

        if status_value not in {"Approved", "Pending", "Rejected"}:
            status_value = "Approved"

        db.add(
            models.AttendanceRecord(
                employee_id=employee.id,
                date=parsed_date.strftime("%d %b %Y"),
                day=day_value or parsed_date.strftime("%a"),
                check_in=scheduled_check_in,
                check_out=scheduled_check_out,
                clock_in=actual_clock_in,
                clock_out=actual_clock_out,
                meal_break=_to_string(row[meal_break_idx]) if meal_break_idx is not None and meal_break_idx < len(row) else "",
                total_hours=total_hours_value,
                work_hours=work_hours_value,
                work_hours_val=_to_hours_value(work_hours_value),
                overtime=_to_string(row[overtime_idx]) if overtime_idx is not None and overtime_idx < len(row) else "",
                approval_status=status_value,
            )
        )
        inserted_count += 1

    db.commit()

    return {
        "employeeId": employee.id,
        "employeeName": employee.name,
        "fromDate": from_date.isoformat(),
        "toDate": to_date.isoformat(),
        "inserted": inserted_count,
        "replaced": replaced_count,
        "skipped": skipped_count,
    }


def _get_attendance_main_module(db: Session, user: models.User) -> dict[str, Any]:
    employee_ids = [employee.id for employee in _get_scoped_employee_query(db, user).all()]
    if not employee_ids:
        monthly = []
    else:
        monthly = (
            db.query(models.MonthlyAttendance)
            .filter(models.MonthlyAttendance.employee_id.in_(employee_ids))
            .all()
        )

    total = len(monthly)
    today_absent = 0
    today_present = 0

    for row in monthly:
        first_non_off = next((d for d in row.days if d in ("P", "A")), None)
        if first_non_off == "A":
            today_absent += 1
        elif first_non_off == "P":
            today_present += 1

    reports = [
        {"id": 1, "title": "Total Employee", "value": total, "icon": "users", "color": "info"},
        {"id": 2, "title": "Absent Employee (Today)", "value": today_absent, "icon": "absent", "color": "danger"},
        {"id": 3, "title": "Present Employee (Today)", "value": today_present, "icon": "present", "color": "success"},
        {"id": 4, "title": "Working Days (Current Month)", "value": 22, "icon": "days", "color": "primary"},
    ]

    table_rows = [
        {"employeeName": row.employee_name, "days": row.days}
        for row in monthly
    ]

    return {"reports": reports, "tableRows": table_rows}


def _get_performance_reviews_module(db: Session, user: models.User) -> dict[str, Any]:
    query = db.query(models.PerformanceReview)
    if not _can_view_cross_department(user):
        visible_names = {employee.name for employee in _get_scoped_employee_query(db, user).all()}
        if not visible_names:
            return {"reviews": []}
        query = query.filter(models.PerformanceReview.employee.in_(visible_names))

    return {
        "reviews": [
            {
                "employee": r.employee,
                "reviewer": r.reviewer,
                "period": r.period,
                "score": r.score,
                "outcome": r.outcome,
            }
            for r in query.all()
        ]
    }


def _get_monthly_attendance_summary(
    db: Session,
    employee_id: int,
    year: int,
    month: int,
) -> dict[str, Any]:
    month_key = f"{year:04d}-{month:02d}"
    monthly_row = (
        db.query(models.MonthlyAttendance)
        .filter(
            models.MonthlyAttendance.employee_id == employee_id,
            models.MonthlyAttendance.month_year == month_key,
        )
        .first()
    )

    if monthly_row and isinstance(monthly_row.days, list):
        tracked_days = sum(1 for value in monthly_row.days if value in {"P", "A"})
        present_days = sum(1 for value in monthly_row.days if value == "P")
        percentage = round((present_days / tracked_days) * 100, 1) if tracked_days else None
        return {
            "percentage": percentage,
            "presentDays": present_days,
            "trackedDays": tracked_days,
            "monthYear": month_key,
        }

    month_records = []
    for record in (
        db.query(models.AttendanceRecord)
        .filter(models.AttendanceRecord.employee_id == employee_id)
        .all()
    ):
        parsed_date = _parse_holiday_date(record.date)
        if parsed_date and parsed_date.year == year and parsed_date.month == month:
            month_records.append(record)

    tracked_days = len(month_records)
    present_days = sum(1 for record in month_records if (record.approval_status or "Approved") != "Rejected")
    percentage = round((present_days / tracked_days) * 100, 1) if tracked_days else None
    return {
        "percentage": percentage,
        "presentDays": present_days,
        "trackedDays": tracked_days,
        "monthYear": month_key,
    }


def _get_review_score_breakdown(db: Session, user: models.User) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    submissions = (
        db.query(models.ReviewSubmission)
        .filter(models.ReviewSubmission.employee_user_id == user.id)
        .order_by(models.ReviewSubmission.id.desc())
        .all()
    )

    review_scores: list[dict[str, Any]] = []
    scored_values: list[float] = []

    for submission in submissions:
        cycle = submission.cycle
        score = float(submission.hr_score) if submission.hr_score is not None else None
        if score is not None:
            scored_values.append(score)

        review_scores.append(
            {
                "id": submission.id,
                "title": cycle.title if cycle and cycle.title else "Performance Review",
                "department": cycle.department if cycle and cycle.department else (submission.department or ""),
                "deadline": cycle.deadline if cycle and cycle.deadline else "",
                "submittedAt": submission.submitted_at or "",
                "status": submission.status or "submitted",
                "score": round(score, 1) if score is not None else None,
                "scorePercentage": round((score / 5) * 100, 1) if score is not None else 0,
                "managerComment": submission.manager_comment or "",
                "hrNotes": submission.hr_notes or "",
                "managerAssessedAt": submission.manager_assessed_at or "",
                "hrScoredAt": submission.hr_scored_at or "",
            }
        )

    latest_scored = next((review for review in review_scores if review["score"] is not None), None)
    average_score = round(sum(scored_values) / len(scored_values), 1) if scored_values else None

    return review_scores, {
        "averageScore": average_score,
        "reviewCount": len(review_scores),
        "scoredCount": len(scored_values),
        "latestScore": latest_scored["score"] if latest_scored else None,
        "latestTitle": latest_scored["title"] if latest_scored else "",
    }


def _get_my_profile_module(db: Session, user: models.User, year: int | None = None, month: int | None = None) -> dict[str, Any]:
    review_scores, performance_summary = _get_review_score_breakdown(db, user)
    employee = _get_employee_for_user(db, user)
    if not employee:
        return {
            "currentMonth": "",
            "selectedDates": [],
            "attendanceRecords": [],
            "performanceMetrics": [],
            "monthlyAttendance": {
                "percentage": None,
                "presentDays": 0,
                "trackedDays": 0,
                "monthYear": "",
            },
            "performanceSummary": performance_summary,
            "reviewScores": review_scores,
            "documents": [],
            "calendarDays": [],
        }

    attendance = (
        db.query(models.AttendanceRecord)
        .filter(models.AttendanceRecord.employee_id == employee.id)
        .all()
    )

    attendance_records = [
        {
            "date": a.date,
            "day": a.day,
            "time": a.check_in,
            "status": "On-Time" if a.approval_status == "Approved" else a.approval_status,
            "statusColor": "text-emerald-600" if a.approval_status == "Approved" else "text-warning",
        }
        for a in attendance
    ]

    metrics = [
        {"label": m.label, "score": m.score, "percentage": m.percentage}
        for m in db.query(models.PerformanceMetric).filter(models.PerformanceMetric.employee_id == employee.id).all()
    ]

    documents = [
        {"name": d.name, "size": d.size, "icon": d.icon}
        for d in db.query(models.Document).filter(models.Document.employee_id == employee.id).all()
    ]

    current_day = date.today()
    current_year = year if year is not None else current_day.year
    current_month = month if month is not None else current_day.month
    monthly_attendance = _get_monthly_attendance_summary(db, employee.id, current_year, current_month)
    days_in_month = month_calendar.monthrange(current_year, current_month)[1]
    first_weekday = date(current_year, current_month, 1).weekday()

    calendar_days: list[dict[str, Any]] = []

    if first_weekday > 0:
        if current_month == 1:
            prev_month = 12
            prev_year = current_year - 1
        else:
            prev_month = current_month - 1
            prev_year = current_year
        prev_month_days = month_calendar.monthrange(prev_year, prev_month)[1]
        for day_value in range(prev_month_days - first_weekday + 1, prev_month_days + 1):
            calendar_days.append({"date": day_value, "month": "prev"})

    holiday_map: dict[int, list[str]] = {}
    for holiday in db.query(models.Holiday).all():
        parsed = _parse_holiday_date(holiday.date)
        if parsed and parsed.year == current_year and parsed.month == current_month:
            holiday_map.setdefault(parsed.day, []).append(holiday.name)

    event_map: dict[int, list[str]] = {}
    month_start_iso = date(current_year, current_month, 1).isoformat()
    month_end_iso = date(current_year, current_month, days_in_month).isoformat()
    for event in db.query(models.CalendarEvent).all():
        if not event.start_date:
            continue
        end_d_str = event.end_date or event.start_date
        try:
            ev_from = date.fromisoformat(event.start_date)
            ev_to = date.fromisoformat(end_d_str)
        except ValueError:
            continue
        if ev_to.isoformat() < month_start_iso or ev_from.isoformat() > month_end_iso:
            continue
        for d in _date_range_days(ev_from, ev_to, current_year, current_month):
            event_map.setdefault(d, []).append(event.title)

    # Get all approved leave applications for the whole company
    leave_map: dict[int, list[str]] = {}
    all_approved_leaves = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.status == "Approved",
            models.LeaveApplication.from_date <= month_end_iso,
            models.LeaveApplication.to_date >= month_start_iso,
        )
        .all()
    )
    for leave_app in all_approved_leaves:
        try:
            from_d = date.fromisoformat(leave_app.from_date)
            to_d = date.fromisoformat(leave_app.to_date)
        except (ValueError, TypeError):
            continue
        for d in _date_range_days(from_d, to_d, current_year, current_month):
            desc = f"{leave_app.employee_name} – on leave"
            leave_map.setdefault(d, []).append(desc)

    for day_value in range(1, days_in_month + 1):
        cell: dict[str, Any] = {"date": day_value}
        day_holidays = holiday_map.get(day_value, [])
        day_events = event_map.get(day_value, [])
        day_leaves = leave_map.get(day_value, [])

        # Combine all events for the cell, removing duplicates
        combined_events = list(day_events)
        for l_desc in day_leaves:
            if l_desc not in combined_events:
                combined_events.append(l_desc)

        if day_holidays:
            cell["bgColor"] = "bg-danger/10"
            cell["holidayNames"] = day_holidays
            if combined_events:
                cell["events"] = combined_events
        elif combined_events:
            # Check if any event is a leave event
            is_leave = any("on leave" in ev.lower() for ev in combined_events)
            cell["bgColor"] = "bg-warning/10" if is_leave else "bg-primary/10"
            cell["events"] = combined_events

        calendar_days.append(cell)


    while len(calendar_days) % 7 != 0:
        calendar_days.append({"date": len(calendar_days) % 7 + 1, "month": "next"})

    leave_balance = get_user_leave_balance_summary(db, user)

    line_manager = "N/A"
    if employee.reporting:
        if isinstance(employee.reporting, list) and len(employee.reporting) > 0:
            line_manager = employee.reporting[0]
        else:
            line_manager = str(employee.reporting)

    return {
        "currentMonth": date(current_year, current_month, 1).strftime("%B %Y"),
        "selectedDates": [9, 10, 11],
        "attendanceRecords": attendance_records,
        "performanceMetrics": metrics,
        "monthlyAttendance": monthly_attendance,
        "performanceSummary": performance_summary,
        "reviewScores": review_scores,
        "documents": documents,
        "calendarDays": calendar_days,
        "leaveBalance": leave_balance,
        "employeeId": employee.employee_code or "",
        "reportingManager": line_manager,
        "dateOfJoining": employee.date_of_joining or "",
        "email": employee.work_email or user.email,
    }


def _get_my_bio_module(db: Session, user: models.User) -> dict[str, Any]:
    ensure_employee_extended_fields(db)
    employee = db.query(models.Employee).filter(models.Employee.user_id == user.id).first()
    if not employee:
        return {}

    dept_map = _department_map(db)
    return {
        "firstName": employee.first_name or '',
        "lastName": employee.last_name or '',
        "gender": employee.gender or '',
        "dateOfBirth": employee.date_of_birth,
        "nationality": employee.nationality or '',
        "maritalStatus": employee.marital_status or '',
        "profilePhoto": employee.profile_photo or '',
        "nationalId": employee.national_id or '',
        "personalEmail": employee.personal_email or '',
        "workEmail": employee.work_email or '',
        "phone": employee.contact or '',
        "emergencyContactName": employee.emergency_contact_name or '',
        "emergencyContactPhone": employee.emergency_contact_phone or '',
        "emergencyContactRelationship": employee.emergency_contact_relationship or '',
        "addressCity": employee.address_city or '',
        "addressDistrict": employee.address_district or '',
        "addressCountry": employee.address_country or '',
        "addressLine1": employee.address_line1 or '',
        "employeeId": employee.employee_code or '',
        "department": dept_map.get(employee.department_id, ''),
        "jobTitle": employee.role or '',
        "reportingManager": '',
        "employmentType": employee.employment_type or '',
        "dateOfJoining": employee.date_of_joining,
        "workLocation": employee.work_location or '',
        "status": employee.status or 'Active',
        "salary": employee.salary or '',
        "payGrade": employee.pay_grade or '',
        "salaryBenefits": employee.salary_benefits or '',
        "bankAccount": employee.bank_account or '',
        "accountNames": employee.account_names or '',
        "bankName": employee.bank_name or '',
        "bankDetails": employee.bank_details or '',
        "taxId": employee.tax_id or '',
        "nssfNumber": employee.nssf_number or '',
        "cvDocument": employee.cv_document or '',
        "contractDocument": employee.contract_document or '',
        "idCopyDocument": employee.id_copy_document or '',
        "certificatesDocument": employee.certificates_document or '',
        "otherDocuments": employee.other_documents or '',
    }


def _get_leave_requests_module(db: Session, user: models.User) -> dict[str, Any]:
    query = db.query(models.LeaveApplication)

    if _can_view_cross_department(user):
        leave_rows = query.all()
    elif _is_staff_role(user.role):
        leave_rows = query.filter(
            (models.LeaveApplication.submitted_by_user_id == user.id)
            | (func.lower(models.LeaveApplication.employee_name) == (user.full_name or "").strip().lower())
        ).all()
    else:
        leave_rows = query.filter(
            func.lower(models.LeaveApplication.department) == (user.department or "").strip().lower()
        ).all()

    approved = sum(1 for l in leave_rows if l.status == "Approved")
    today_leaves = sum(1 for l in leave_rows if l.status == "Pending")

    return {
        "summary": [
            {
                "id": 1,
                "title": "Today/Presents Leave",
                "value": f"{approved}/{max(len(leave_rows), 1)}",
                "description": "Today/Presents Leave",
                "icon": "present",
                "textColor": "text-primary",
                "bgColor": "bg-primary/10",
            },
            {
                "id": 2,
                "title": "Today Leaves",
                "value": str(today_leaves),
                "description": "Today Leaves",
                "icon": "today",
                "textColor": "text-success",
                "bgColor": "bg-success/10",
            },
        ],
        "rows": [
            {
                "id": row.id,
                "employeeName": row.employee_name or "N/A",
                "leaveType": row.leave_type,
                "reason": row.reason,
                "noOfDays": row.days,
                "from": row.from_date,
                "to": row.to_date,
                "status": row.status,
            }
            for row in leave_rows
        ],
    }


def _get_calendar_module(db: Session) -> dict[str, Any]:
    external_events = [
        {"title": e.title, "className": e.class_name}
        for e in db.query(models.ExternalEvent).all()
    ]
    events = [_serialize_calendar_event(e) for e in db.query(models.CalendarEvent).all()]

    holidays = db.query(models.Holiday).all()
    for holiday in holidays:
        holiday_event = _serialize_holiday_event(holiday)
        if holiday_event:
            events.append(holiday_event)

    return {"externalEvents": external_events, "events": events}


# Leave planner / applications endpoints support

def _date_range_days(from_d: date, to_d: date, target_year: int, target_month: int) -> list[int]:
    """Return day-of-month numbers within a date range that fall in the given year/month."""
    result: list[int] = []
    current = from_d
    while current <= to_d:
        if current.year == target_year and current.month == target_month:
            result.append(current.day)
        current += timedelta(days=1)
    return result


def _count_working_days(from_date_str: str, to_date_str: str) -> int:
    """Count weekdays (Mon-Fri) between two ISO date strings, inclusive."""
    try:
        from_d = date.fromisoformat(from_date_str)
        to_d = date.fromisoformat(to_date_str)
    except (ValueError, TypeError):
        return 1
    count = 0
    current = from_d
    while current <= to_d:
        if current.weekday() < 5:
            count += 1
        current += timedelta(days=1)
    return max(count, 1)


def _serialize_leave_app(row: models.LeaveApplication) -> dict[str, Any]:
    handover_due_date = None
    try:
        if row.from_date:
            handover_due_date = (date.fromisoformat(row.from_date) - timedelta(days=1)).isoformat()
    except (TypeError, ValueError):
        handover_due_date = None

    return {
        "id": row.id,
        "leaveType": row.leave_type or "",
        "reason": row.reason or "",
        "days": row.days or 1,
        "fromDate": row.from_date or "",
        "toDate": row.to_date or "",
        # legacy keys kept for the applications table in leave-planner
        "from": row.from_date or "",
        "to": row.to_date or "",
        "reviewer": row.reviewer or "Line Manager",
        "submittedOn": row.submitted_on or "",
        "status": row.status or "Draft",
        "employeeName": row.employee_name or "",
        "department": row.department or "",
        "standIn": row.stand_in or "",
        "submittedByUserId": row.submitted_by_user_id,
        "handoverReport": row.handover_report or "",
        "handoverSubmittedOn": row.handover_submitted_on or "",
        "handoverDueDate": handover_due_date or "",
    }


def get_stand_in_options(db: Session, user: models.User) -> list[dict[str, Any]]:
    """Return all active staff members who can be selected as stand-ins.

    All users can pick from the full active staff list, excluding themselves.
    """
    rows = (
        db.query(models.User)
        .filter(
            models.User.is_active == True,  # noqa: E712
            models.User.id != user.id,
        )
        .order_by(models.User.full_name)
        .all()
    )
    return [
        {
            "id": row.id,
            "fullName": row.full_name,
            "role": row.role or "",
            "department": row.department or "",
        }
        for row in rows
    ]


def get_leave_calendar_cells(db: Session, user: models.User | None = None, year: int | None = None, month: int | None = None) -> list[dict[str, Any]]:
    target = date.today()
    target_year = year or target.year
    target_month = month or target.month

    month_days = month_calendar.monthrange(target_year, target_month)[1]
    first_weekday_index = date(target_year, target_month, 1).weekday()  # Monday=0

    cells: list[dict[str, Any]] = []

    if first_weekday_index > 0:
        if target_month == 1:
            prev_month_year = target_year - 1
            prev_month = 12
        else:
            prev_month_year = target_year
            prev_month = target_month - 1

        prev_month_days = month_calendar.monthrange(prev_month_year, prev_month)[1]
        start_prev_day = prev_month_days - first_weekday_index + 1
        for day in range(start_prev_day, prev_month_days + 1):
            cells.append({"day": day, "outside": True, "muted": True})

    for day in range(1, month_days + 1):
        weekday = date(target_year, target_month, day).weekday()
        cells.append({"day": day, "weekend": weekday >= 5})

    while len(cells) % 7 != 0:
        cells.append({"day": len(cells) % 7 + 1, "outside": True, "muted": True})

    holiday_map: dict[int, list[str]] = {}
    for holiday in db.query(models.Holiday).all():
        parsed = _parse_holiday_date(holiday.date)
        if not parsed:
            continue
        if parsed.year != target_year or parsed.month != target_month:
            continue
        holiday_map.setdefault(parsed.day, []).append(holiday.name)

    # Calendar events (regular + ranged leave events)
    event_map: dict[int, list[str]] = {}
    leave_names_map: dict[int, list[str]] = {}
    month_start_iso = date(target_year, target_month, 1).isoformat()
    month_end_iso = date(target_year, target_month, month_days).isoformat()
    for event in db.query(models.CalendarEvent).all():
        if not event.start_date:
            continue
        end_d_str = event.end_date or event.start_date
        try:
            ev_from = date.fromisoformat(event.start_date)
            ev_to = date.fromisoformat(end_d_str)
        except ValueError:
            continue
        # skip if the event range doesn't overlap the target month
        if ev_to.isoformat() < month_start_iso or ev_from.isoformat() > month_end_iso:
            continue
        for d in _date_range_days(ev_from, ev_to, target_year, target_month):
            event_map.setdefault(d, []).append(event.title)
            is_leave = (event.event_type == "leave") or ("on leave" in (event.title or "").lower())
            if is_leave:
                name = ""
                if event.employee_name:
                    name = event.employee_name.strip()
                else:
                    title = event.title or ""
                    if " – on leave" in title:
                        name = title.split(" – on leave")[0].strip()
                    elif " on leave" in title.lower():
                        idx = title.lower().find(" on leave")
                        name = title[:idx].strip()
                if name:
                    first_name = name.split()[0]
                    if first_name not in leave_names_map.setdefault(d, []):
                        leave_names_map[d].append(first_name)

    # User's own Draft / Pending leave — block those days for themselves
    own_draft_days: set[int] = set()
    own_pending_days: set[int] = set()
    # Same-department approved leave by colleagues — shown as "team away"
    team_away_map: dict[int, list[str]] = {}  # day -> [employee names]

    if user:
        dept = (user.department or "").strip().lower()

        own_leaves = (
            db.query(models.LeaveApplication)
            .filter(
                models.LeaveApplication.submitted_by_user_id == user.id,
                models.LeaveApplication.status.in_(["Draft", "Pending"]),
                models.LeaveApplication.from_date <= month_end_iso,
                models.LeaveApplication.to_date >= month_start_iso,
            )
            .all()
        )
        for leave in own_leaves:
            try:
                lv_from = date.fromisoformat(leave.from_date)
                lv_to = date.fromisoformat(leave.to_date)
            except (ValueError, TypeError):
                continue
            for d in _date_range_days(lv_from, lv_to, target_year, target_month):
                if leave.status == "Draft":
                    own_draft_days.add(d)
                else:
                    own_pending_days.add(d)

        if dept:
            dept_colleague_leaves = (
                db.query(models.LeaveApplication)
                .filter(
                    models.LeaveApplication.submitted_by_user_id != user.id,
                    models.LeaveApplication.status.in_(["Draft", "Pending", "Approved"]),
                    func.lower(models.LeaveApplication.department) == dept,
                    models.LeaveApplication.from_date <= month_end_iso,
                    models.LeaveApplication.to_date >= month_start_iso,
                )
                .all()
            )
            for leave in dept_colleague_leaves:
                try:
                    lv_from = date.fromisoformat(leave.from_date)
                    lv_to = date.fromisoformat(leave.to_date)
                except (ValueError, TypeError):
                    continue
                name = (leave.employee_name or "A colleague").split()[0]
                for d in _date_range_days(lv_from, lv_to, target_year, target_month):
                    team_away_map.setdefault(d, []).append(name)

    for cell in cells:
        if cell.get("outside"):
            continue

        day = cell.get("day")
        if not isinstance(day, int):
            continue

        day_holidays = holiday_map.get(day, [])
        day_events = event_map.get(day, [])

        # Collect unique leave first names from CalendarEvents and team_away_map
        day_leave_names = list(leave_names_map.get(day, []))
        if day in team_away_map:
            for name in team_away_map[day]:
                if name not in day_leave_names:
                    day_leave_names.append(name)

        if day_holidays:
            cell["unavailable"] = True
            cell["badge"] = "HOLIDAY"
            cell["holidayNames"] = day_holidays

        if day_leave_names:
            cell.setdefault("events", [])
            for name in day_leave_names:
                desc = f"{name} – on leave"
                if desc not in cell["events"]:
                    cell["events"].append(desc)
            if "badge" not in cell:
                cell["badge"] = ", ".join(day_leave_names)
            cell["unavailable"] = True

        if day_events:
            cell.setdefault("events", [])
            for ev in day_events:
                if ev not in cell["events"]:
                    cell["events"].append(ev)
            if "badge" not in cell:
                cell["badge"] = "EVENT"

        if day in own_draft_days:
            cell["draftBlock"] = True
            cell["unavailable"] = True

        if day in own_pending_days:
            cell["pendingBlock"] = True
            cell["unavailable"] = True

        if day in team_away_map:
            cell["teamAway"] = True
            cell["unavailable"] = True

    return cells


def get_leave_applications(db: Session, user: models.User) -> list[dict[str, Any]]:
    query = db.query(models.LeaveApplication).filter(
        models.LeaveApplication.status != "Draft"
    )
    if _can_view_cross_department(user):
        rows = query.order_by(models.LeaveApplication.id.desc()).all()
    elif _is_staff_role(user.role):
        rows = (
            query
            .filter(
                (models.LeaveApplication.submitted_by_user_id == user.id)
                | (func.lower(models.LeaveApplication.employee_name) == (user.full_name or "").strip().lower())
            )
            .order_by(models.LeaveApplication.id.desc())
            .all()
        )
    else:
        rows = (
            query
            .filter(func.lower(models.LeaveApplication.department) == (user.department or "").strip().lower())
            .order_by(models.LeaveApplication.id.desc())
            .all()
        )

    return [_serialize_leave_app(row) for row in rows]


def create_leave_application(db: Session, payload: dict[str, Any], user: models.User) -> dict[str, Any]:
    requested_days = _count_working_days(payload["from"], payload["to"])
    balance = get_user_leave_balance_summary(db, user)
    
    l_type = (payload.get("leaveType") or "").strip().lower()
    if "sick" in l_type:
        remaining = balance["sick"]["remainingDays"]
    elif "emergency" in l_type:
        remaining = balance["emergency"]["remainingDays"]
    elif "unpaid" in l_type:
        remaining = 999  # Unpaid leave is not limited by balance
    else:
        remaining = balance["annual"]["remainingDays"]
        
    if requested_days > remaining:
        return f"insufficient_balance:{requested_days}:{remaining}"


    row = models.LeaveApplication(
        employee_name=payload.get("employeeName") or user.full_name,
        leave_type=payload["leaveType"],
        reason=payload["reason"],
        days=requested_days,
        from_date=payload["from"],
        to_date=payload["to"],
        reviewer=payload["reviewer"],
        submitted_on=date.today().strftime("%d %b, %Y"),
        department=user.department,
        submitted_by_user_id=user.id,
        status="Pending",
    )
    db.add(row)
    db.commit()
    db.refresh(row)

    # Send leave request notifications
    try:
        reviewer_emails = []
        reviewer_users = []
        
        r_user = _find_active_user_by_full_name(db, row.reviewer)
        if r_user:
            reviewer_users.append(r_user)
            if r_user.email:
                reviewer_emails.append(r_user.email)
        
        # Fallback to employee's reporting managers if no specific reviewer matched
        if not reviewer_users:
            employee = db.query(models.Employee).filter(models.Employee.user_id == user.id).first()
            if employee and employee.reporting:
                for mgr_name in employee.reporting:
                    mgr_user = _find_active_user_by_full_name(db, mgr_name)
                    if mgr_user and mgr_user not in reviewer_users:
                        reviewer_users.append(mgr_user)
                        if mgr_user.email:
                            reviewer_emails.append(mgr_user.email)
                            
        # Add in-app notifications
        add_notification(
            db,
            user_id=user.id,
            title="Leave Request Submitted",
            message=f"Your {row.leave_type} request for {row.from_date} to {row.to_date} ({row.days} day(s)) has been submitted.",
            type="leave"
        )
        for r_usr in reviewer_users:
            add_notification(
                db,
                user_id=r_usr.id,
                title="Pending Leave Request",
                message=f"{row.employee_name} requested {row.leave_type} from {row.from_date} to {row.to_date} ({row.days} day(s)).",
                type="leave"
            )
            
        from app.email import send_leave_request_notification
        # Send confirmation to employee
        send_leave_request_notification(
            recipient_email=user.email,
            requester_name=row.employee_name,
            leave_type=row.leave_type,
            start_date=row.from_date,
            end_date=row.to_date,
            days=row.days,
            reason=row.reason,
            reviewer_name=row.reviewer,
            is_reviewer=False
        )
        # Send notification to resolved reviewers
        for r_email in reviewer_emails:
            send_leave_request_notification(
                recipient_email=r_email,
                requester_name=row.employee_name,
                leave_type=row.leave_type,
                start_date=row.from_date,
                end_date=row.to_date,
                days=row.days,
                reason=row.reason,
                reviewer_name=row.reviewer,
                is_reviewer=True
            )
    except Exception as e:
        logger.error(f"Failed to send leave request email: {e}")

    return {
        "id": row.id,
        "leaveType": row.leave_type,
        "reason": row.reason,
        "days": row.days,
        "from": row.from_date,
        "to": row.to_date,
        "reviewer": row.reviewer,
        "submittedOn": row.submitted_on,
        "status": row.status,
    }


def cancel_leave(db: Session, leave_id: int, user: models.User) -> dict[str, Any] | None:
    row = db.query(models.LeaveApplication).filter(models.LeaveApplication.id == leave_id).first()
    if not row:
        return None

    if not _can_view_cross_department(user):
        if _is_staff_role(user.role):
            owned_by_user = row.submitted_by_user_id == user.id or (
                (row.employee_name or "").strip().lower() == (user.full_name or "").strip().lower()
            )
            if not owned_by_user:
                return {}
        elif (row.department or "").strip().lower() != (user.department or "").strip().lower():
            return {}

    row.status = "Declined"
    db.commit()
    db.refresh(row)
    return {
        "id": row.id,
        "leaveType": row.leave_type,
        "reason": row.reason,
        "days": row.days,
        "from": row.from_date,
        "to": row.to_date,
        "reviewer": row.reviewer,
        "submittedOn": row.submitted_on,
        "status": row.status,
    }


def create_draft_leave(db: Session, payload: dict[str, Any], user: models.User) -> dict[str, Any] | str:
    from_date = payload["fromDate"]
    to_date = payload["toDate"]
    leave_type = payload.get("leaveType") or "Annual Leave"
    days = _count_working_days(from_date, to_date)

    dept = (user.department or "").strip().lower()
    if dept:
        conflict = (
            db.query(models.LeaveApplication)
            .filter(
                models.LeaveApplication.submitted_by_user_id != user.id,
                func.lower(models.LeaveApplication.department) == dept,
                models.LeaveApplication.status.in_(["Draft", "Pending", "Approved"]),
                models.LeaveApplication.from_date <= to_date,
                models.LeaveApplication.to_date >= from_date,
            )
            .first()
        )
        if conflict:
            return f"conflict:{conflict.employee_name}:{conflict.from_date}:{conflict.to_date}"

    row = models.LeaveApplication(
        employee_name=user.full_name,
        leave_type=leave_type,
        reason="",
        days=days,
        from_date=from_date,
        to_date=to_date,
        reviewer="Line Manager",
        submitted_on=date.today().isoformat(),
        department=user.department,
        submitted_by_user_id=user.id,
        status="Draft",
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _serialize_leave_app(row)


def get_draft_leaves(db: Session, user: models.User) -> list[dict[str, Any]]:
    rows = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.submitted_by_user_id == user.id,
            models.LeaveApplication.status == "Draft",
        )
        .order_by(models.LeaveApplication.id.asc())
        .all()
    )
    return [_serialize_leave_app(r) for r in rows]


def delete_draft_leave(db: Session, leave_id: int, user: models.User) -> bool:
    row = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.id == leave_id,
            models.LeaveApplication.submitted_by_user_id == user.id,
            models.LeaveApplication.status == "Draft",
        )
        .first()
    )
    if not row:
        return False
    db.delete(row)
    db.commit()
    return True


def submit_leave(db: Session, leave_id: int, payload: dict[str, Any], user: models.User) -> dict[str, Any] | str:
    """Transition a Draft leave to Pending after checking for same-department conflicts.

    Returns the updated row dict on success, or a string starting with 'conflict:' or 'not_found'.
    """
    row = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.id == leave_id,
            models.LeaveApplication.submitted_by_user_id == user.id,
            models.LeaveApplication.status == "Draft",
        )
        .first()
    )
    if not row:
        return "not_found"

    if payload.get("leaveType"):
        row.leave_type = payload["leaveType"]
    if payload.get("reason"):
        row.reason = payload["reason"]
    if payload.get("standIn") is not None:
        row.stand_in = payload["standIn"] or None

    balance = get_user_leave_balance_summary(db, user)
    
    l_type = (row.leave_type or "").strip().lower()
    if "sick" in l_type:
        remaining = balance["sick"]["remainingDays"]
    elif "emergency" in l_type:
        remaining = balance["emergency"]["remainingDays"]
    elif "unpaid" in l_type:
        remaining = 999  # Unpaid leave is not limited by balance
    else:
        remaining = balance["annual"]["remainingDays"]
        
    if int(row.days or 0) > remaining:
        return f"insufficient_balance:{int(row.days or 0)}:{remaining}"


    dept = (user.department or "").strip().lower()
    if dept:
        conflict = (
            db.query(models.LeaveApplication)
            .filter(
                models.LeaveApplication.id != row.id,
                models.LeaveApplication.submitted_by_user_id != user.id,
                func.lower(models.LeaveApplication.department) == dept,
                models.LeaveApplication.status.in_(["Draft", "Pending", "Approved"]),
                models.LeaveApplication.from_date <= row.to_date,
                models.LeaveApplication.to_date >= row.from_date,
            )
            .first()
        )
        if conflict:
            return f"conflict:{conflict.employee_name}:{conflict.from_date}:{conflict.to_date}"

    row.status = "Pending"
    row.submitted_on = date.today().isoformat()
    db.commit()
    db.refresh(row)

    # Send leave request notifications
    try:
        reviewer_emails = []
        reviewer_users = []
        
        r_user = _find_active_user_by_full_name(db, row.reviewer)
        if r_user:
            reviewer_users.append(r_user)
            if r_user.email:
                reviewer_emails.append(r_user.email)
        
        # Fallback to employee's reporting managers if no specific reviewer matched
        if not reviewer_users:
            employee = db.query(models.Employee).filter(models.Employee.user_id == user.id).first()
            if employee and employee.reporting:
                for mgr_name in employee.reporting:
                    mgr_user = _find_active_user_by_full_name(db, mgr_name)
                    if mgr_user and mgr_user not in reviewer_users:
                        reviewer_users.append(mgr_user)
                        if mgr_user.email:
                            reviewer_emails.append(mgr_user.email)
                            
        # Add in-app notifications
        add_notification(
            db,
            user_id=user.id,
            title="Leave Request Submitted",
            message=f"Your {row.leave_type} request for {row.from_date} to {row.to_date} ({row.days} day(s)) has been submitted.",
            type="leave"
        )
        for r_usr in reviewer_users:
            add_notification(
                db,
                user_id=r_usr.id,
                title="Pending Leave Request",
                message=f"{row.employee_name} requested {row.leave_type} from {row.from_date} to {row.to_date} ({row.days} day(s)).",
                type="leave"
            )
        
        from app.email import send_leave_request_notification
        # Send confirmation to employee
        send_leave_request_notification(
            recipient_email=user.email,
            requester_name=row.employee_name,
            leave_type=row.leave_type,
            start_date=row.from_date,
            end_date=row.to_date,
            days=row.days,
            reason=row.reason,
            reviewer_name=row.reviewer,
            is_reviewer=False
        )
        # Send notification to resolved reviewers
        for r_email in reviewer_emails:
            send_leave_request_notification(
                recipient_email=r_email,
                requester_name=row.employee_name,
                leave_type=row.leave_type,
                start_date=row.from_date,
                end_date=row.to_date,
                days=row.days,
                reason=row.reason,
                reviewer_name=row.reviewer,
                is_reviewer=True
            )
    except Exception as e:
        logger.error(f"Failed to send leave submission email: {e}")

    return _serialize_leave_app(row)


def update_pending_leave(db: Session, leave_id: int, payload: dict[str, Any], user: models.User) -> dict[str, Any] | str:
    """Update a Pending leave application (reason, leaveType, standIn)."""
    row = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.id == leave_id,
            models.LeaveApplication.submitted_by_user_id == user.id,
            models.LeaveApplication.status == "Pending",
        )
        .first()
    )
    if not row:
        return "not_found"

    if payload.get("leaveType"):
        row.leave_type = payload["leaveType"]
    if "reason" in payload:
        row.reason = payload["reason"]
    if "standIn" in payload:
        row.stand_in = payload["standIn"] or None

    db.commit()
    db.refresh(row)
    return _serialize_leave_app(row)


def delete_pending_leave(db: Session, leave_id: int, user: models.User) -> bool:
    """Delete a Pending leave application owned by the user."""
    row = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.id == leave_id,
            models.LeaveApplication.submitted_by_user_id == user.id,
            models.LeaveApplication.status == "Pending",
        )
        .first()
    )
    if not row:
        return False
    db.delete(row)
    db.commit()
    return True


def get_marked_dates(db: Session) -> list[dict[str, Any]]:
    rows = db.query(models.MarkedDate).all()
    return [
        {
            "id": row.id,
            "dates": row.dates,
            "markedOn": row.marked_on,
            "status": row.status,
        }
        for row in rows
    ]


def apply_marked_dates(db: Session, marked_id: int) -> dict[str, Any] | None:
    row = db.query(models.MarkedDate).filter(models.MarkedDate.id == marked_id).first()
    if not row:
        return None
    row.status = "Applied"
    db.commit()
    db.refresh(row)
    return {
        "id": row.id,
        "dates": row.dates,
        "markedOn": row.marked_on,
        "status": row.status,
    }


def get_approvals(db: Session, user: models.User | None = None) -> list[dict[str, Any]]:
    query = db.query(models.LeaveApplication).filter(
        models.LeaveApplication.status == "Pending"
    )
    # Non-HR managers only see their own department
    if user and not _can_view_cross_department(user):
        dept = (user.department or "").strip().lower()
        query = query.filter(func.lower(models.LeaveApplication.department) == dept)
    rows = query.order_by(models.LeaveApplication.id.desc()).all()
    return [
        {
            "id": row.id,
            "employee": row.employee_name or "",
            "department": row.department or "",
            "leaveType": row.leave_type or "",
            "period": f"{row.from_date} → {row.to_date}",
            "fromDate": row.from_date or "",
            "toDate": row.to_date or "",
            "days": row.days or 1,
            "approver": row.reviewer or "Line Manager",
            "submittedOn": row.submitted_on or "",
            "status": row.status,
            "standIn": row.stand_in or "",
            "handoverReport": row.handover_report or "",
            "handoverSubmittedOn": row.handover_submitted_on or "",
        }
        for row in rows
    ]


def _find_active_user_by_full_name(db: Session, full_name: str | None) -> models.User | None:
    name = (full_name or "").strip().lower()
    if not name:
        return None
    
    # Try exact match first
    user = (
        db.query(models.User)
        .filter(
            models.User.is_active == True,  # noqa: E712
            func.lower(models.User.full_name) == name,
        )
        .first()
    )
    if user:
        return user
        
    # Fallback to order-independent matching of words
    words = [w for w in name.split() if w]
    if not words:
        return None
        
    all_users = db.query(models.User).filter(models.User.is_active == True).all()  # noqa: E712
    for u in all_users:
        u_name = (u.full_name or "").strip().lower()
        if not u_name:
            continue
        u_words = [w for w in u_name.split() if w]
        if sorted(words) == sorted(u_words):
            return u
            
    return None



def _send_handover_email(
    *,
    recipient_email: str | None,
    recipient_name: str | None,
    leave_owner: str | None,
    leave_from: str | None,
    leave_to: str | None,
    report: str,
) -> bool:
    smtp_host = (os.getenv("HR_TOOL_SMTP_HOST") or "").strip()
    smtp_username = (os.getenv("HR_TOOL_SMTP_USERNAME") or "").strip()
    smtp_password = os.getenv("HR_TOOL_SMTP_PASSWORD") or ""
    sender = (os.getenv("HR_TOOL_SMTP_SENDER") or smtp_username).strip()

    if not smtp_host or not sender or not recipient_email:
        return False

    smtp_port_raw = (os.getenv("HR_TOOL_SMTP_PORT") or "587").strip()
    try:
        smtp_port = int(smtp_port_raw)
    except ValueError:
        smtp_port = 587

    use_tls = (os.getenv("HR_TOOL_SMTP_USE_TLS") or "true").strip().lower() in {"1", "true", "yes", "on"}

    subject = f"Hand-over report from {leave_owner or 'colleague'}"
    body = (
        f"Hello {recipient_name or 'Colleague'},\n\n"
        f"{leave_owner or 'A colleague'} has shared a hand-over report for leave coverage.\n"
        f"Leave period: {leave_from or 'N/A'} to {leave_to or 'N/A'}.\n\n"
        "Hand-over report:\n"
        f"{report}\n\n"
        "Please review and follow up where needed."
    )

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = recipient_email
    msg.set_content(body)

    try:
        if use_tls:
            context = ssl.create_default_context()
            with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as smtp:
                smtp.starttls(context=context)
                if smtp_username and smtp_password:
                    smtp.login(smtp_username, smtp_password)
                smtp.send_message(msg)
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as smtp:
                if smtp_username and smtp_password:
                    smtp.login(smtp_username, smtp_password)
                smtp.send_message(msg)
    except Exception:
        logger.exception("Failed to send hand-over email")
        return False

    return True


def submit_handover_report(
    db: Session,
    leave_id: int,
    report: str,
    user: models.User,
) -> dict[str, Any] | str:
    row = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.id == leave_id,
            models.LeaveApplication.submitted_by_user_id == user.id,
            models.LeaveApplication.status == "Approved",
        )
        .first()
    )
    if not row:
        return "not_found"

    report_text = (report or "").strip()
    if not report_text:
        return "invalid_report"

    row.handover_report = report_text
    row.handover_submitted_on = date.today().isoformat()
    db.commit()
    db.refresh(row)

    stand_in_user = _find_active_user_by_full_name(db, row.stand_in)
    email_sent = _send_handover_email(
        recipient_email=stand_in_user.email if stand_in_user else None,
        recipient_name=stand_in_user.full_name if stand_in_user else row.stand_in,
        leave_owner=row.employee_name,
        leave_from=row.from_date,
        leave_to=row.to_date,
        report=report_text,
    )

    payload = _serialize_leave_app(row)
    payload["handoverEmailSent"] = email_sent
    return payload


def action_approval(db: Session, approval_id: int, action: str) -> dict[str, Any] | None:
    row = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.id == approval_id,
            models.LeaveApplication.status == "Pending",
        )
        .first()
    )
    if not row:
        return None

    if action not in ("Approved", "Rejected"):
        return None

    row.status = action
    db.commit()

    if action == "Approved":
        # Create a calendar event spanning the leave date range so all users see it
        event = models.CalendarEvent(
            title=f"{row.employee_name or 'Staff'} – on leave",
            start_date=row.from_date,
            end_date=row.to_date,
            class_names=["!bg-rose-100", "!text-rose-700"],
            employee_name=row.employee_name,
            department=row.department,
            event_type="leave",
            leave_application_id=row.id,
        )
        db.add(event)
        db.commit()

    db.refresh(row)

    # Send leave action notification to the employee
    try:
        employee_email = None
        employee_user = None
        if row.submitted_by_user_id:
            employee_user = db.query(models.User).filter(models.User.id == row.submitted_by_user_id).first()
        if not employee_user and row.employee_name:
            employee_user = _find_active_user_by_full_name(db, row.employee_name)
        
        if employee_user:
            employee_email = employee_user.email
            # Add in-app notification
            add_notification(
                db,
                user_id=employee_user.id,
                title=f"Leave Request {row.status}",
                message=f"Your {row.leave_type} request for {row.from_date} to {row.to_date} ({row.days} day(s)) was {row.status.lower()} by {row.reviewer or 'Line Manager'}.",
                type="leave"
            )
        
        if employee_email:
            from app.email import send_leave_action_notification
            send_leave_action_notification(
                recipient_email=employee_email,
                requester_name=row.employee_name,
                leave_type=row.leave_type,
                start_date=row.from_date,
                end_date=row.to_date,
                days=row.days,
                status=row.status,
                actioned_by=row.reviewer or "Line Manager"
            )
    except Exception as e:
        logger.error(f"Failed to send leave action email: {e}")

    return {
        "id": row.id,
        "employee": row.employee_name or "",
        "department": row.department or "",
        "leaveType": row.leave_type or "",
        "period": f"{row.from_date} → {row.to_date}",
        "days": row.days or 1,
        "approver": row.reviewer or "Line Manager",
        "status": row.status,
    }


def get_user_by_email(db: Session, email: str) -> models.User | None:
    return db.query(models.User).filter(models.User.email == email).first()


def get_role_access_profiles(db: Session) -> list[dict[str, Any]]:
    rows = (
        db.query(models.RoleAccessProfile)
        .order_by(models.RoleAccessProfile.is_system.desc(), models.RoleAccessProfile.name.asc())
        .all()
    )
    return [
        _serialize_role_access_profile(row)
        for row in rows
    ]


def _serialize_role_access_profile(row: models.RoleAccessProfile) -> dict[str, Any]:
    return {
        "id": row.id,
        "name": row.name,
        "description": row.description or "",
        "accessLevel": row.access_level,
        "dashboardRoute": row.dashboard_route,
        "permissions": row.permissions or [],
        "isSystem": bool(row.is_system),
    }


def get_role_access_profile_options(db: Session) -> list[dict[str, str]]:
    rows = (
        db.query(models.RoleAccessProfile)
        .order_by(models.RoleAccessProfile.is_system.desc(), models.RoleAccessProfile.name.asc())
        .all()
    )
    return [
        {
            "name": row.name,
            "accessLevel": row.access_level,
            "dashboardRoute": row.dashboard_route,
        }
        for row in rows
        if (row.name or "").strip().lower() != "super-admin"
    ]


def create_role_access_profile(db: Session, payload: dict[str, Any]) -> dict[str, Any] | None:
    existing = (
        db.query(models.RoleAccessProfile)
        .filter(func.lower(models.RoleAccessProfile.name) == payload["name"].strip().lower())
        .first()
    )
    if existing:
        return None

    row = models.RoleAccessProfile(
        name=payload["name"].strip(),
        description=payload.get("description", "").strip(),
        access_level=payload["accessLevel"].strip(),
        dashboard_route=payload["dashboardRoute"].strip(),
        permissions=payload.get("permissions", []),
        is_system=False,
    )
    db.add(row)
    db.commit()
    db.refresh(row)

    return _serialize_role_access_profile(row)


def update_role_access_profile(db: Session, role_profile_id: int, payload: dict[str, Any]) -> dict[str, Any] | None:
    row = db.query(models.RoleAccessProfile).filter(models.RoleAccessProfile.id == role_profile_id).first()
    if not row:
        return None

    if row.is_system:
        return {"error": "system-profile-immutable"}

    next_name = payload["name"].strip()
    duplicate = (
        db.query(models.RoleAccessProfile)
        .filter(
            func.lower(models.RoleAccessProfile.name) == next_name.lower(),
            models.RoleAccessProfile.id != role_profile_id,
        )
        .first()
    )
    if duplicate:
        return {}

    old_name = (row.name or "").strip()
    row.name = next_name
    row.description = payload.get("description", "").strip()
    row.access_level = payload["accessLevel"].strip()
    row.dashboard_route = payload["dashboardRoute"].strip()
    row.permissions = payload.get("permissions", [])

    role_sections = _extract_section_keys_from_permissions(row.permissions or [])
    old_name_lower = old_name.lower()

    linked_users = db.query(models.User).filter(func.lower(models.User.role) == old_name_lower).all()
    for user in linked_users:
        user.role = row.name
        user.access_level = row.access_level
        user.dashboard_route = row.dashboard_route
        if (user.role or "").strip().lower() == "super-admin":
            user.allowed_sections = ["*"]
        else:
            user.allowed_sections = role_sections

    db.commit()
    db.refresh(row)

    payload = _serialize_role_access_profile(row)
    payload["linkedUsersUpdated"] = len(linked_users)
    return payload


def delete_role_access_profile(db: Session, role_profile_id: int) -> dict[str, Any] | None:
    row = db.query(models.RoleAccessProfile).filter(models.RoleAccessProfile.id == role_profile_id).first()
    if not row:
        return None

    if row.is_system:
        return {"error": "system-profile-immutable"}

    role_name_lower = (row.name or "").strip().lower()
    linked_users_count = db.query(models.User).filter(func.lower(models.User.role) == role_name_lower).count()
    if linked_users_count > 0:
        return {
            "error": "role-in-use",
            "usersCount": linked_users_count,
        }

    deleted_profile = _serialize_role_access_profile(row)
    db.delete(row)
    db.commit()

    return {
        "deleted": True,
        "profile": deleted_profile,
    }


def _serialize_department(row: models.Department, initial_leave_days: int | None = None) -> dict[str, Any]:
    return {
        "id": row.id,
        "name": row.name,
        "head": row.head or "",
        "lineManager": row.line_manager or "",
        "hrManager": row.hr_manager or "",
        "staffCount": int(row.staff_count or 0),
        "approvalLevel": row.approval_level or "",
        "status": row.status or "Active",
        "initialLeaveDays": int(initial_leave_days or DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS),
        "icon": row.icon or "solar:buildings-3-bold-duotone",
        "iconBg": row.icon_bg or "bg-primary/10 text-primary",
        "avatarColor": row.avatar_color or "bg-slate-200 dark:bg-slate-700",
    }


def _ensure_department_leave_policy_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS department_leave_policies (
                department_id INTEGER PRIMARY KEY,
                initial_leave_days INTEGER NOT NULL DEFAULT 21,
                updated_by_user_id INTEGER,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
    )


def _get_department_leave_policy_map_for_ids(
    db: Session,
    department_ids: list[int],
) -> dict[int, int]:
    if not department_ids:
        return {}

    _ensure_department_leave_policy_table(db)
    placeholders = ", ".join(str(department_id) for department_id in department_ids)
    rows = db.execute(
        text(
            f"SELECT department_id, initial_leave_days FROM department_leave_policies "
            f"WHERE department_id IN ({placeholders})"
        )
    ).fetchall()
    return {
        int(row[0]): max(int(row[1] or DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS), 1)
        for row in rows
    }


def _get_department_leave_policy_map(db: Session) -> dict[int, int]:
    department_ids = [row[0] for row in db.query(models.Department.id).all()]
    return _get_department_leave_policy_map_for_ids(db, [int(value) for value in department_ids])


def _sanitize_initial_leave_days(value: Any) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS
    return max(parsed, 1)


def _set_department_initial_leave_days(
    db: Session,
    department_id: int,
    initial_leave_days: int,
    updated_by_user_id: int | None = None,
) -> None:
    _ensure_department_leave_policy_table(db)

    existing_row = db.execute(
        text(
            """
            SELECT initial_leave_days
            FROM department_leave_policies
            WHERE department_id = :department_id
            """
        ),
        {"department_id": department_id},
    ).first()
    previous_days = int(existing_row[0]) if existing_row else None
    next_days = max(initial_leave_days, 1)

    db.execute(
        text(
            """
            INSERT INTO department_leave_policies (department_id, initial_leave_days, updated_by_user_id, updated_at)
            VALUES (:department_id, :initial_leave_days, :updated_by_user_id, CURRENT_TIMESTAMP)
            ON CONFLICT(department_id) DO UPDATE SET
                initial_leave_days = excluded.initial_leave_days,
                updated_by_user_id = excluded.updated_by_user_id,
                updated_at = CURRENT_TIMESTAMP
            """
        ),
        {
            "department_id": department_id,
            "initial_leave_days": next_days,
            "updated_by_user_id": updated_by_user_id,
        },
    )

    if previous_days != next_days:
        _add_department_leave_policy_audit(
            db,
            department_id=department_id,
            previous_days=previous_days,
            new_days=next_days,
            changed_by_user_id=updated_by_user_id,
        )


def _get_department_initial_leave_days_by_name(db: Session, department_name: str | None) -> int:
    normalized_name = (department_name or "").strip().lower()
    if not normalized_name:
        return DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS

    department = (
        db.query(models.Department)
        .filter(func.lower(models.Department.name) == normalized_name)
        .first()
    )
    if not department:
        return DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS

    policy_map = _get_department_leave_policy_map_for_ids(db, [department.id])
    return int(policy_map.get(department.id, DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS))


def get_user_leave_balance_summary(db: Session, user: models.User) -> dict[str, Any]:
    # Initial days
    initial_annual = _get_department_initial_leave_days_by_name(db, user.department)
    initial_sick = 10
    initial_emergency = 5
    initial_unpaid = 0

    # Retrieve all APPROVED leave applications for this user
    leaves = (
        db.query(models.LeaveApplication)
        .filter(
            models.LeaveApplication.submitted_by_user_id == user.id,
            models.LeaveApplication.status == "Approved",
        )
        .all()
    )

    used_annual = 0
    used_sick = 0
    used_emergency = 0
    used_unpaid = 0

    for l in leaves:
        l_type = (l.leave_type or "").strip().lower()
        if "sick" in l_type:
            used_sick += int(l.days or 0)
        elif "emergency" in l_type:
            used_emergency += int(l.days or 0)
        elif "unpaid" in l_type:
            used_unpaid += int(l.days or 0)
        else:
            # Defaults to Annual
            used_annual += int(l.days or 0)

    remaining_annual = max(initial_annual - used_annual, 0)
    remaining_sick = max(initial_sick - used_sick, 0)
    remaining_emergency = max(initial_emergency - used_emergency, 0)

    return {
        "annual": {
            "initialDays": initial_annual,
            "usedDays": used_annual,
            "remainingDays": remaining_annual,
        },
        "sick": {
            "initialDays": initial_sick,
            "usedDays": used_sick,
            "remainingDays": remaining_sick,
        },
        "emergency": {
            "initialDays": initial_emergency,
            "usedDays": used_emergency,
            "remainingDays": remaining_emergency,
        },
        "unpaid": {
            "initialDays": initial_unpaid,
            "usedDays": used_unpaid,
            "remainingDays": 0,
        },
        # Root level keys for backward compatibility (Annual leave details)
        "initialDays": initial_annual,
        "usedDays": used_annual,
        "remainingDays": remaining_annual,
    }



def list_departments(db: Session) -> list[dict[str, Any]]:
    rows = db.query(models.Department).order_by(models.Department.name.asc()).all()
    leave_policy_map = _get_department_leave_policy_map(db)
    result: list[dict[str, Any]] = []
    for row in rows:
        payload = {
            "id": row.id,
            "name": row.name,
            "head": row.head or "",
            "lineManager": row.line_manager or "",
            "hrManager": row.hr_manager or "",
            "staffCount": int(row.staff_count or 0),
            "approvalLevel": row.approval_level or "",
            "status": row.status or "Active",
            "initialLeaveDays": int(leave_policy_map.get(row.id, DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS)),
            "icon": row.icon or "solar:buildings-3-bold-duotone",
            "iconBg": row.icon_bg or "bg-primary/10 text-primary",
            "avatarColor": row.avatar_color or "bg-slate-200 dark:bg-slate-700",
        }
        result.append(payload)
    return result


def get_department_by_name(db: Session, name: str) -> models.Department | None:
    return db.query(models.Department).filter(func.lower(models.Department.name) == name.strip().lower()).first()


def create_department(db: Session, payload: dict[str, Any]) -> dict[str, Any] | None:
    if get_department_by_name(db, payload["name"]):
        return None

    row = models.Department(
        name=payload["name"].strip(),
        head=payload.get("head", "").strip() or None,
        line_manager=payload.get("lineManager", "").strip() or None,
        hr_manager=payload.get("hrManager", "").strip() or None,
        staff_count=max(int(payload.get("staffCount", 0)), 0),
        approval_level=payload.get("approvalLevel", "1-Step Approval").strip() or "1-Step Approval",
        status=payload.get("status", "Active").strip() or "Active",
        icon=payload.get("icon", "solar:buildings-3-bold-duotone"),
        icon_bg=payload.get("iconBg", "bg-primary/10 text-primary"),
        avatar_color=payload.get("avatarColor", "bg-slate-200 dark:bg-slate-700"),
    )
    db.add(row)
    db.commit()
    db.refresh(row)

    _set_department_initial_leave_days(
        db,
        row.id,
        _sanitize_initial_leave_days(payload.get("initialLeaveDays", DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS)),
        payload.get("updatedByUserId"),
    )
    db.commit()

    leave_policy_map = _get_department_leave_policy_map_for_ids(db, [row.id])
    return {
        "id": row.id,
        "name": row.name,
        "head": row.head or "",
        "lineManager": row.line_manager or "",
        "hrManager": row.hr_manager or "",
        "staffCount": int(row.staff_count or 0),
        "approvalLevel": row.approval_level or "",
        "status": row.status or "Active",
        "initialLeaveDays": int(leave_policy_map.get(row.id, DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS)),
        "icon": row.icon or "solar:buildings-3-bold-duotone",
        "iconBg": row.icon_bg or "bg-primary/10 text-primary",
        "avatarColor": row.avatar_color or "bg-slate-200 dark:bg-slate-700",
    }


def update_department(db: Session, department_id: int, payload: dict[str, Any]) -> dict[str, Any] | None:
    row = db.query(models.Department).filter(models.Department.id == department_id).first()
    if not row:
        return None

    incoming_name = payload.get("name", row.name).strip()
    duplicate = (
        db.query(models.Department)
        .filter(
            func.lower(models.Department.name) == incoming_name.lower(),
            models.Department.id != department_id,
        )
        .first()
    )
    if duplicate:
        return {}

    row.name = incoming_name
    row.head = payload.get("head", row.head)
    row.line_manager = payload.get("lineManager", row.line_manager)
    row.hr_manager = payload.get("hrManager", row.hr_manager)
    row.staff_count = max(int(payload.get("staffCount", row.staff_count or 0)), 0)
    row.approval_level = payload.get("approvalLevel", row.approval_level or "1-Step Approval")
    row.status = payload.get("status", row.status or "Active")
    row.icon = payload.get("icon", row.icon)
    row.icon_bg = payload.get("iconBg", row.icon_bg)
    row.avatar_color = payload.get("avatarColor", row.avatar_color)

    if "initialLeaveDays" in payload:
        _set_department_initial_leave_days(
            db,
            row.id,
            _sanitize_initial_leave_days(payload.get("initialLeaveDays")),
            payload.get("updatedByUserId"),
        )

    db.commit()
    db.refresh(row)
    leave_policy_map = _get_department_leave_policy_map_for_ids(db, [row.id])
    return {
        "id": row.id,
        "name": row.name,
        "head": row.head or "",
        "lineManager": row.line_manager or "",
        "hrManager": row.hr_manager or "",
        "staffCount": int(row.staff_count or 0),
        "approvalLevel": row.approval_level or "",
        "status": row.status or "Active",
        "initialLeaveDays": int(leave_policy_map.get(row.id, DEFAULT_DEPARTMENT_INITIAL_LEAVE_DAYS)),
        "icon": row.icon or "solar:buildings-3-bold-duotone",
        "iconBg": row.icon_bg or "bg-primary/10 text-primary",
        "avatarColor": row.avatar_color or "bg-slate-200 dark:bg-slate-700",
    }


def delete_department(db: Session, department_id: int) -> bool:
    row = db.query(models.Department).filter(models.Department.id == department_id).first()
    if not row:
        return False

    db.delete(row)
    db.commit()
    return True


def get_role_profile_by_name(db: Session, role_profile_name: str) -> models.RoleAccessProfile | None:
    return (
        db.query(models.RoleAccessProfile)
        .filter(func.lower(models.RoleAccessProfile.name) == role_profile_name.strip().lower())
        .first()
    )


def _serialize_user(row: models.User) -> dict[str, Any]:
    normalized_role = (row.role or "").strip().lower()
    allowed_sections = ["*"] if normalized_role == "super-admin" else (row.allowed_sections or [])

    return {
        "id": row.id,
        "email": row.email,
        "fullName": row.full_name,
        "department": row.department or "General",
        "role": row.role,
        "accessLevel": row.access_level or "self-service",
        "dashboardRoute": resolve_dashboard_route(
            row.role,
            row.department,
            row.access_level,
            row.dashboard_route,
        ),
        "allowedSections": allowed_sections,
        "isActive": bool(row.is_active),
    }


def _ensure_user_audit_log_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS user_audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                action TEXT NOT NULL,
                actor_user_id INTEGER,
                actor_email TEXT,
                target_user_id INTEGER,
                target_email TEXT,
                details TEXT
            )
            """
        )
    )


def add_user_audit_log(
    db: Session,
    *,
    action: str,
    actor_user_id: int | None,
    actor_email: str | None,
    target_user_id: int | None,
    target_email: str | None,
    details: dict[str, Any] | None = None,
) -> None:
    _ensure_user_audit_log_table(db)
    db.execute(
        text(
            """
            INSERT INTO user_audit_logs (
                action,
                actor_user_id,
                actor_email,
                target_user_id,
                target_email,
                details
            ) VALUES (
                :action,
                :actor_user_id,
                :actor_email,
                :target_user_id,
                :target_email,
                :details
            )
            """
        ),
        {
            "action": action,
            "actor_user_id": actor_user_id,
            "actor_email": actor_email,
            "target_user_id": target_user_id,
            "target_email": target_email,
            "details": json.dumps(details or {}),
        },
    )


def _ensure_notification_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                message TEXT NOT NULL,
                type TEXT DEFAULT 'general',
                is_read BOOLEAN DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
    )


def add_notification(db: Session, user_id: int, title: str, message: str, type: str = "general") -> None:
    _ensure_notification_table(db)
    db.execute(
        text(
            """
            INSERT INTO notifications (user_id, title, message, type)
            VALUES (:user_id, :title, :message, :type)
            """
        ),
        {
            "user_id": user_id,
            "title": title,
            "message": message,
            "type": type,
        }
    )
    db.commit()


def list_notifications(db: Session, user_id: int) -> list[dict[str, Any]]:
    _ensure_notification_table(db)
    rows = db.execute(
        text(
            """
            SELECT id, user_id, title, message, type, is_read, created_at
            FROM notifications
            WHERE user_id = :user_id
            ORDER BY id DESC
            """
        ),
        {"user_id": user_id}
    ).fetchall()
    
    return [
        {
            "id": row._mapping.get("id"),
            "userId": row._mapping.get("user_id"),
            "title": row._mapping.get("title"),
            "message": row._mapping.get("message"),
            "type": row._mapping.get("type"),
            "isRead": bool(row._mapping.get("is_read")),
            "createdAt": row._mapping.get("created_at")
        }
        for row in rows
    ]


def mark_notification_as_read(db: Session, notification_id: int, user_id: int) -> bool:
    _ensure_notification_table(db)
    result = db.execute(
        text(
            """
            UPDATE notifications
            SET is_read = 1
            WHERE id = :id AND user_id = :user_id
            """
        ),
        {"id": notification_id, "user_id": user_id}
    )
    db.commit()
    return result.rowcount > 0


def mark_all_notifications_as_read(db: Session, user_id: int) -> None:
    _ensure_notification_table(db)
    db.execute(
        text(
            """
            UPDATE notifications
            SET is_read = 1
            WHERE user_id = :user_id
            """
        ),
        {"user_id": user_id}
    )
    db.commit()


def _ensure_profile_update_request_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS profile_update_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                requester_user_id INTEGER NOT NULL,
                requester_name TEXT,
                requester_department TEXT,
                requested_fields TEXT,
                note TEXT,
                status TEXT NOT NULL DEFAULT 'Pending',
                reviewed_by_user_id INTEGER,
                reviewed_by_name TEXT,
                reviewed_at TEXT,
                review_note TEXT
            )
            """
        )
    )

    existing_columns = {
        row[1]
        for row in db.execute(text("PRAGMA table_info(profile_update_requests)")).fetchall()
    }

    alter_statements = {
        "reviewed_by_user_id": "ALTER TABLE profile_update_requests ADD COLUMN reviewed_by_user_id INTEGER",
        "reviewed_by_name": "ALTER TABLE profile_update_requests ADD COLUMN reviewed_by_name TEXT",
        "reviewed_at": "ALTER TABLE profile_update_requests ADD COLUMN reviewed_at TEXT",
        "review_note": "ALTER TABLE profile_update_requests ADD COLUMN review_note TEXT",
    }
    for column_name, statement in alter_statements.items():
        if column_name not in existing_columns:
            db.execute(text(statement))


def create_profile_update_request(
    db: Session,
    *,
    requester_user_id: int,
    requester_name: str,
    requester_department: str,
    requested_fields: list[str],
    note: str,
) -> dict[str, Any]:
    _ensure_profile_update_request_table(db)
    inserted = db.execute(
        text(
            """
            INSERT INTO profile_update_requests (
                requester_user_id,
                requester_name,
                requester_department,
                requested_fields,
                note
            ) VALUES (
                :requester_user_id,
                :requester_name,
                :requester_department,
                :requested_fields,
                :note
            )
            """
        ),
        {
            "requester_user_id": requester_user_id,
            "requester_name": requester_name,
            "requester_department": requester_department,
            "requested_fields": json.dumps(requested_fields),
            "note": note,
        },
    )
    db.commit()
    return {
        "id": int(inserted.lastrowid or 0),
        "message": "Your request has been submitted and HR will review it shortly.",
        "requestedFields": requested_fields,
    }


def list_profile_update_requests(
    db: Session,
    *,
    status: str | None = None,
) -> list[dict[str, Any]]:
    _ensure_profile_update_request_table(db)
    base_sql = (
        "SELECT id, created_at, requester_user_id, requester_name, requester_department, "
        "requested_fields, note, status, reviewed_by_user_id, reviewed_by_name, reviewed_at, review_note "
        "FROM profile_update_requests"
    )

    params: dict[str, Any] = {}
    if status and status.strip().lower() != "all":
        base_sql += " WHERE lower(status) = :status"
        params["status"] = status.strip().lower()

    base_sql += " ORDER BY id DESC"
    rows = db.execute(text(base_sql), params).fetchall()

    items: list[dict[str, Any]] = []
    for row in rows:
        payload = dict(row._mapping)
        raw_fields = payload.get("requested_fields")
        requested_fields: list[str] = []
        if isinstance(raw_fields, str):
            try:
                parsed_fields = json.loads(raw_fields)
                if isinstance(parsed_fields, list):
                    requested_fields = [str(value) for value in parsed_fields]
            except json.JSONDecodeError:
                requested_fields = []

        items.append(
            {
                "id": payload.get("id"),
                "createdAt": payload.get("created_at") or "",
                "requesterUserId": payload.get("requester_user_id"),
                "requesterName": payload.get("requester_name") or "",
                "requesterDepartment": payload.get("requester_department") or "",
                "requestedFields": requested_fields,
                "note": payload.get("note") or "",
                "status": payload.get("status") or "Pending",
                "reviewedByUserId": payload.get("reviewed_by_user_id"),
                "reviewedByName": payload.get("reviewed_by_name") or "",
                "reviewedAt": payload.get("reviewed_at") or "",
                "reviewNote": payload.get("review_note") or "",
            }
        )

    return items


def process_profile_update_request(
    db: Session,
    *,
    request_id: int,
    action: str,
    reviewer_user_id: int,
    reviewer_name: str,
    review_note: str = "",
) -> dict[str, Any] | None:
    _ensure_profile_update_request_table(db)

    row = db.execute(
        text(
            """
            SELECT id, status
            FROM profile_update_requests
            WHERE id = :request_id
            """
        ),
        {"request_id": request_id},
    ).first()
    if not row:
        return None

    normalized_action = (action or "").strip().lower()
    next_status_map = {
        "approve": "Approved",
        "approved": "Approved",
        "reject": "Rejected",
        "rejected": "Rejected",
        "in-progress": "In Progress",
        "in_progress": "In Progress",
        "pending": "Pending",
    }
    next_status = next_status_map.get(normalized_action)
    if not next_status:
        return {}

    db.execute(
        text(
            """
            UPDATE profile_update_requests
            SET status = :status,
                reviewed_by_user_id = :reviewed_by_user_id,
                reviewed_by_name = :reviewed_by_name,
                reviewed_at = CURRENT_TIMESTAMP,
                review_note = :review_note
            WHERE id = :request_id
            """
        ),
        {
            "status": next_status,
            "reviewed_by_user_id": reviewer_user_id,
            "reviewed_by_name": reviewer_name,
            "review_note": (review_note or "").strip(),
            "request_id": request_id,
        },
    )
    db.commit()

    updated = db.execute(
        text(
            """
            SELECT id, created_at, requester_user_id, requester_name, requester_department,
                   requested_fields, note, status, reviewed_by_user_id, reviewed_by_name,
                   reviewed_at, review_note
            FROM profile_update_requests
            WHERE id = :request_id
            """
        ),
        {"request_id": request_id},
    ).first()
    if not updated:
        return None

    payload = dict(updated._mapping)
    raw_fields = payload.get("requested_fields")
    requested_fields: list[str] = []
    if isinstance(raw_fields, str):
        try:
            parsed_fields = json.loads(raw_fields)
            if isinstance(parsed_fields, list):
                requested_fields = [str(value) for value in parsed_fields]
        except json.JSONDecodeError:
            requested_fields = []

    return {
        "id": payload.get("id"),
        "createdAt": payload.get("created_at") or "",
        "requesterUserId": payload.get("requester_user_id"),
        "requesterName": payload.get("requester_name") or "",
        "requesterDepartment": payload.get("requester_department") or "",
        "requestedFields": requested_fields,
        "note": payload.get("note") or "",
        "status": payload.get("status") or "Pending",
        "reviewedByUserId": payload.get("reviewed_by_user_id"),
        "reviewedByName": payload.get("reviewed_by_name") or "",
        "reviewedAt": payload.get("reviewed_at") or "",
        "reviewNote": payload.get("review_note") or "",
    }


def get_profile_update_request_employee_context(
    db: Session,
    *,
    request_id: int,
) -> dict[str, Any] | None:
    _ensure_profile_update_request_table(db)

    row = db.execute(
        text(
            """
            SELECT id, requester_user_id, requester_name, requester_department, requested_fields, note, status
            FROM profile_update_requests
            WHERE id = :request_id
            """
        ),
        {"request_id": request_id},
    ).first()
    if not row:
        return None

    payload = dict(row._mapping)
    requester_user_id = payload.get("requester_user_id")
    employee = None
    if requester_user_id:
        employee = db.query(models.Employee).filter(models.Employee.user_id == int(requester_user_id)).first()

    requested_fields: list[str] = []
    raw_fields = payload.get("requested_fields")
    if isinstance(raw_fields, str):
        try:
            parsed = json.loads(raw_fields)
            if isinstance(parsed, list):
                requested_fields = [str(value) for value in parsed]
        except json.JSONDecodeError:
            requested_fields = []

    employee_payload: dict[str, Any] | None = None
    if employee:
        employee_payload = {
            "employeeDbId": employee.id,
            "name": employee.name or "",
            "phone": employee.contact or "",
            "personalEmail": employee.personal_email or "",
            "workEmail": employee.work_email or "",
            "emergencyContactName": employee.emergency_contact_name or "",
            "emergencyContactPhone": employee.emergency_contact_phone or "",
            "emergencyContactRelationship": employee.emergency_contact_relationship or "",
            "addressCity": employee.address_city or "",
            "addressDistrict": employee.address_district or "",
            "addressCountry": employee.address_country or "",
            "addressLine1": employee.address_line1 or "",
            "maritalStatus": employee.marital_status or "",
            "nationality": employee.nationality or "",
            "nationalId": employee.national_id or "",
            "workLocation": employee.work_location or "",
            "salary": employee.salary or "",
            "payGrade": employee.pay_grade or "",
            "salaryBenefits": employee.salary_benefits or "",
            "bankAccount": employee.bank_account or "",
            "accountNames": employee.account_names or "",
            "bankName": employee.bank_name or "",
            "bankDetails": employee.bank_details or "",
            "taxId": employee.tax_id or "",
            "nssfNumber": employee.nssf_number or "",
        }

    return {
        "request": {
            "id": payload.get("id"),
            "requesterUserId": requester_user_id,
            "requesterName": payload.get("requester_name") or "",
            "requesterDepartment": payload.get("requester_department") or "",
            "requestedFields": requested_fields,
            "note": payload.get("note") or "",
            "status": payload.get("status") or "Pending",
        },
        "employee": employee_payload,
    }


def _ensure_department_leave_policy_audit_table(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS department_leave_policy_audits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                department_id INTEGER NOT NULL,
                department_name TEXT,
                previous_days INTEGER,
                new_days INTEGER NOT NULL,
                changed_by_user_id INTEGER,
                changed_by_name TEXT
            )
            """
        )
    )


def _add_department_leave_policy_audit(
    db: Session,
    *,
    department_id: int,
    previous_days: int | None,
    new_days: int,
    changed_by_user_id: int | None,
) -> None:
    _ensure_department_leave_policy_audit_table(db)
    department = db.query(models.Department).filter(models.Department.id == department_id).first()
    changed_by_name = "System"
    if changed_by_user_id:
        actor = db.query(models.User).filter(models.User.id == changed_by_user_id).first()
        if actor and actor.full_name:
            changed_by_name = actor.full_name

    db.execute(
        text(
            """
            INSERT INTO department_leave_policy_audits (
                department_id,
                department_name,
                previous_days,
                new_days,
                changed_by_user_id,
                changed_by_name
            ) VALUES (
                :department_id,
                :department_name,
                :previous_days,
                :new_days,
                :changed_by_user_id,
                :changed_by_name
            )
            """
        ),
        {
            "department_id": department_id,
            "department_name": (department.name if department else ""),
            "previous_days": previous_days,
            "new_days": new_days,
            "changed_by_user_id": changed_by_user_id,
            "changed_by_name": changed_by_name,
        },
    )


def list_department_leave_policy_audits(
    db: Session,
    *,
    department_id: int | None = None,
    limit: int = 200,
) -> list[dict[str, Any]]:
    _ensure_department_leave_policy_audit_table(db)

    base_sql = (
        "SELECT id, created_at, department_id, department_name, previous_days, new_days, "
        "changed_by_user_id, changed_by_name FROM department_leave_policy_audits"
    )
    params: dict[str, Any] = {"limit_value": int(limit)}

    if department_id is not None:
        base_sql += " WHERE department_id = :department_id"
        params["department_id"] = int(department_id)

    base_sql += " ORDER BY id DESC LIMIT :limit_value"
    rows = db.execute(text(base_sql), params).fetchall()
    return [
        {
            "id": row._mapping.get("id"),
            "createdAt": row._mapping.get("created_at") or "",
            "departmentId": row._mapping.get("department_id"),
            "departmentName": row._mapping.get("department_name") or "",
            "previousDays": row._mapping.get("previous_days"),
            "newDays": row._mapping.get("new_days"),
            "changedByUserId": row._mapping.get("changed_by_user_id"),
            "changedByName": row._mapping.get("changed_by_name") or "",
        }
        for row in rows
    ]


def list_user_audit_logs(db: Session, target_user_id: int, limit: int = 10) -> list[dict[str, Any]]:
    _ensure_user_audit_log_table(db)
    rows = db.execute(
        text(
            """
            SELECT id, created_at, action, actor_user_id, actor_email, target_user_id, target_email, details
            FROM user_audit_logs
            WHERE target_user_id = :target_user_id
            ORDER BY id DESC
            LIMIT :limit_value
            """
        ),
        {
            "target_user_id": target_user_id,
            "limit_value": int(limit),
        },
    ).all()

    results: list[dict[str, Any]] = []
    for row in rows:
        payload = dict(row._mapping)
        raw_details = payload.get("details")
        if isinstance(raw_details, str):
            try:
                payload["details"] = json.loads(raw_details)
            except json.JSONDecodeError:
                payload["details"] = {"raw": raw_details}
        results.append(payload)
    return results


def list_users(
    db: Session,
    *,
    q: str | None = None,
    status: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> dict[str, Any]:
    query = db.query(models.User)

    normalized_query = (q or "").strip().lower()
    if normalized_query:
        like_value = f"%{normalized_query}%"
        query = query.filter(
            func.lower(models.User.full_name).like(like_value)
            | func.lower(models.User.email).like(like_value)
            | func.lower(models.User.department).like(like_value)
            | func.lower(models.User.role).like(like_value)
        )

    normalized_status = (status or "").strip().lower()
    if normalized_status == "verified":
        query = query.filter(models.User.is_active.is_(True))
    elif normalized_status == "rejected":
        query = query.filter(models.User.is_active.is_(False))

    safe_page = max(1, int(page))
    safe_page_size = min(100, max(1, int(page_size)))

    total = query.count()
    rows = (
        query.order_by(models.User.id.asc())
        .offset((safe_page - 1) * safe_page_size)
        .limit(safe_page_size)
        .all()
    )

    return {
        "items": [_serialize_user(row) for row in rows],
        "total": total,
        "page": safe_page,
        "pageSize": safe_page_size,
    }


def get_user_by_id(db: Session, user_id: int) -> dict[str, Any] | None:
    row = db.query(models.User).filter(models.User.id == user_id).first()
    if not row:
        return None

    payload = _serialize_user(row)
    linked_employee = db.query(models.Employee).filter(models.Employee.user_id == row.id).first()
    payload["employeeId"] = linked_employee.id if linked_employee else None
    payload["employeeName"] = linked_employee.name if linked_employee else None
    payload["linkedToEmployee"] = linked_employee is not None
    payload["recentAuditLogs"] = list_user_audit_logs(db, user_id, limit=10)
    return payload


def update_user(db: Session, user_id: int, payload: dict[str, Any]) -> dict[str, Any] | None:
    row = db.query(models.User).filter(models.User.id == user_id).first()
    if not row:
        return None

    next_email = payload["email"].strip().lower()
    existing = (
        db.query(models.User)
        .filter(func.lower(models.User.email) == next_email, models.User.id != user_id)
        .first()
    )
    if existing:
        return {}

    role_profile_name = payload["role"].strip()
    role_profile = get_role_profile_by_name(db, role_profile_name)
    if not role_profile:
        return {"error": "role-profile-not-found"}

    row.email = next_email
    row.full_name = payload["fullName"].strip()
    row.department = payload["department"].strip() or "General"
    row.role = role_profile.name
    row.access_level = role_profile.access_level
    row.dashboard_route = role_profile.dashboard_route or resolve_dashboard_route(
        row.role,
        row.department,
        row.access_level,
        None,
    )
    row.allowed_sections = _extract_section_keys_from_permissions(role_profile.permissions or [])

    if (row.role or "").strip().lower() == "super-admin":
        row.allowed_sections = ["*"]

    row.is_active = bool(payload.get("isActive", True))

    db.commit()
    db.refresh(row)
    return _serialize_user(row)


def delete_user(db: Session, user_id: int) -> bool | None:
    row = db.query(models.User).filter(models.User.id == user_id).first()
    if not row:
        return None

    # Preserve dependent rows while removing the user account.
    db.query(models.Employee).filter(models.Employee.user_id == user_id).update({models.Employee.user_id: None})
    db.query(models.LeaveApplication).filter(models.LeaveApplication.submitted_by_user_id == user_id).update(
        {models.LeaveApplication.submitted_by_user_id: None}
    )

    db.delete(row)
    db.commit()
    return True


def delete_employee(db: Session, employee_id: int) -> bool | None:
    row = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not row:
        return None

    linked_user_id = row.user_id
    db.delete(row)
    db.flush()

    if linked_user_id:
        user = db.query(models.User).filter(models.User.id == linked_user_id).first()
        if user:
            db.query(models.LeaveApplication).filter(
                models.LeaveApplication.submitted_by_user_id == linked_user_id
            ).update({models.LeaveApplication.submitted_by_user_id: None})
            db.delete(user)

    db.commit()
    return True


def create_user_with_role_profile(db: Session, payload: dict[str, Any]) -> dict[str, Any] | None:
    existing = db.query(models.User).filter(func.lower(models.User.email) == payload["email"].strip().lower()).first()
    if existing:
        return None

    role_profile_name = payload.get("roleProfileName", "").strip()
    role_profile = None
    if role_profile_name:
        role_profile = get_role_profile_by_name(db, role_profile_name)
        if not role_profile:
            return {}

    role = payload.get("role", "").strip() or (role_profile.name if role_profile else "staff")
    access_level = payload.get("accessLevel", "").strip() or (role_profile.access_level if role_profile else "self-service")
    dashboard_route = payload.get("dashboardRoute", "").strip() or (
        role_profile.dashboard_route if role_profile else resolve_dashboard_route(role, payload.get("department"), access_level)
    )
    allowed_sections = payload.get("allowedSections")
    if allowed_sections is None:
        # Extract section keys from granular permissions in the role profile
        allowed_sections = _extract_section_keys_from_permissions(role_profile.permissions) if role_profile else []

    if role.strip().lower() == "super-admin":
        allowed_sections = ["*"]

    row = models.User(
        email=payload["email"].strip().lower(),
        password_hash=payload["password"],
        full_name=payload["fullName"].strip(),
        department=payload.get("department", "General").strip() or "General",
        role=role,
        access_level=access_level,
        dashboard_route=dashboard_route,
        allowed_sections=allowed_sections,
        is_active=bool(payload.get("isActive", True)),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return _serialize_user(row)


def _build_initials(full_name: str) -> str:
    tokens = [part for part in full_name.strip().split(" ") if part]
    if not tokens:
        return "NA"
    if len(tokens) == 1:
        return tokens[0][:2].upper()
    return (tokens[0][0] + tokens[-1][0]).upper()


def create_employee_with_user_account(db: Session, payload: dict[str, Any]) -> dict[str, Any] | None:
    ensure_employee_extended_fields(db)
    department_name = (payload.get("department") or "").strip()
    department = get_department_by_name(db, department_name) if department_name else None
    # Department is optional for top-level / cross-functional roles
    if department_name and not department:
        return {}

    role_profile_name = payload["roleProfileName"].strip()
    role_profile = get_role_profile_by_name(db, role_profile_name)
    if not role_profile:
        return {"error": "role-profile-not-found"}

    if (role_profile.name or "").strip().lower() == "super-admin":
        return {"error": "forbidden-role-profile"}

    normalized_email = payload["email"].strip().lower()
    existing_user = db.query(models.User).filter(func.lower(models.User.email) == normalized_email).first()
    if existing_user:
        return {"error": "email-exists"}

    full_name = payload["fullName"].strip()
    temporary_password = (payload.get("temporaryPassword") or "").strip() or "TempPass@123"

    user_row = models.User(
        email=normalized_email,
        password_hash=temporary_password,
        full_name=full_name,
        department=department.name if department else "",
        role=role_profile.name,
        access_level=role_profile.access_level,
        dashboard_route=role_profile.dashboard_route,
        allowed_sections=_extract_section_keys_from_permissions(role_profile.permissions) or [],
        is_active=True,
    )
    db.add(user_row)
    db.flush()

    reporting = []
    # Filter out sentinel values (__none__ = intentionally no manager, __self__ already resolved on frontend)
    line_manager = (payload.get("lineManager") or "").strip()
    dept_head = (payload.get("departmentHead") or "").strip()
    if line_manager and line_manager != "__none__":
        reporting.append(line_manager)
    if dept_head and dept_head != "__self__":
        reporting.append(dept_head)

    employee_row = models.Employee(
        user_id=user_row.id,
        initials=_build_initials(full_name),
        name=full_name,
        first_name=payload.get("firstName", "").strip(),
        last_name=payload.get("lastName", "").strip(),
        gender=payload.get("gender", "").strip(),
        date_of_birth=payload.get("dateOfBirth", "").strip(),
        nationality=payload.get("nationality", "").strip(),
        marital_status=payload.get("maritalStatus", "").strip(),
        profile_photo=payload.get("profilePhoto", "").strip(),
        national_id=payload.get("nationalId", "").strip(),
        personal_email=payload.get("personalEmail", "").strip(),
        work_email=payload.get("workEmail", "").strip(),
        contact=payload.get("phone", "").strip() or normalized_email,
        emergency_contact_name=payload.get("emergencyContactName", "").strip(),
        emergency_contact_phone=payload.get("emergencyContactPhone", "").strip(),
        emergency_contact_relationship=payload.get("emergencyContactRelationship", "").strip(),
        address_city=payload.get("addressCity", "").strip(),
        address_district=payload.get("addressDistrict", "").strip(),
        address_country=payload.get("addressCountry", "").strip(),
        address_line1=payload.get("addressLine1", "").strip(),
        employee_code=payload.get("employeeId", "").strip(),
        role=payload.get("jobTitle", "").strip() or role_profile.name,
        department_id=department.id if department else None,
        employment_type=payload.get("employmentType", "").strip(),
        date_of_joining=payload.get("dateOfJoining", "").strip(),
        work_location=payload.get("workLocation", "").strip(),
        salary=payload.get("salary", "").strip(),
        pay_grade=payload.get("payGrade", "").strip(),
        salary_benefits=payload.get("salaryBenefits", "").strip(),
        bank_account=payload.get("bankAccount", "").strip(),
        account_names=payload.get("accountNames", "").strip(),
        bank_name=payload.get("bankName", "").strip(),
        bank_details=payload.get("bankDetails", "").strip(),
        tax_id=payload.get("taxId", "").strip(),
        nssf_number=payload.get("nssfNumber", "").strip(),
        cv_document=payload.get("cvDocument", "").strip(),
        contract_document=payload.get("contractDocument", "").strip(),
        id_copy_document=payload.get("idCopyDocument", "").strip(),
        certificates_document=payload.get("certificatesDocument", "").strip(),
        other_documents=payload.get("otherDocuments", "").strip(),
        status=payload.get("status", "Active").strip() or "Active",
        reporting=reporting,
        started_at=payload.get("dateOfJoining", "") or date.today().strftime("%d %b, %Y"),
    )
    db.add(employee_row)

    if department:
        department.staff_count = int(department.staff_count or 0) + 1

    db.commit()
    db.refresh(user_row)
    db.refresh(employee_row)

    # Add welcome in-app notification
    try:
        add_notification(
            db,
            user_id=user_row.id,
            title="Welcome to HR Tool Portal",
            message=f"Hello {user_row.full_name}, your employee profile and account have been set up successfully. Welcome to the team!",
            type="general"
        )
    except Exception as e:
        logger.error(f"Failed to save onboarding in-app notification: {e}")

    # Send onboarding welcome email to the user
    try:
        from app.email import send_welcome_email
        send_welcome_email(
            recipient_email=user_row.email,
            full_name=user_row.full_name,
            temporary_password=temporary_password,
            role=user_row.role,
            department=user_row.department or "General"
        )
    except Exception as e:
        logger.error(f"Failed to trigger onboarding email sending: {e}")

    return {
        "employeeId": employee_row.id,
        "userId": user_row.id,
        "employeeName": employee_row.name,
        "email": user_row.email,
        "role": user_row.role,
        "department": user_row.department,
        "temporaryPassword": temporary_password,
    }
